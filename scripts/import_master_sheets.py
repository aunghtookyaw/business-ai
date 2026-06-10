import argparse
from datetime import datetime, timezone
from pathlib import Path
import sys

import psycopg2.extras
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import config
from tools.formula_engine import _connect
from tools import master_data


DEFAULT_MASTER_DIR = Path("/Users/bigshot/MEGA/Master sheet")


MASTER_SHEETS = {
    "customers": {
        "path": DEFAULT_MASTER_DIR / "customers.xlsx",
        "table": "customer_master",
        "column": "customer_name",
        "expected_header": "Customers",
        "meta_type": "customer",
    },
    "categories": {
        "path": DEFAULT_MASTER_DIR / "categories.xlsx",
        "table": "category_master",
        "column": "category_name",
        "expected_header": "Categorization",
        "meta_type": "category",
    },
}


def _clean_name(value):
    value = str(value or "").strip()
    return value or None


def _read_single_column_workbook(path, expected_header):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = _clean_name(rows[0][0])
    if header and header.lower() != expected_header.lower():
        raise ValueError(f"{path.name} first column must be {expected_header!r}, got {header!r}")

    names = []
    seen = set()
    for row_number, row in enumerate(rows[1:], start=2):
        name = _clean_name(row[0] if row else None)
        normalized = master_data.normalize_name(name)
        if not name or not normalized or normalized in seen:
            continue
        seen.add(normalized)
        names.append({
            "name": name,
            "normalized": normalized,
            "row_number": row_number,
            "sheet": worksheet.title,
        })
    return names


def _metadata(sheet_config, row, imported_at):
    return {
        "source": "mega_master_sheet",
        "source_file": str(sheet_config["path"]),
        "source_sheet": row["sheet"],
        "source_row": row["row_number"],
        "master_type": sheet_config["meta_type"],
        "normalized_name": row["normalized"],
        "search_tokens": row["normalized"].split(),
        "aliases": [],
        "active": True,
        "imported_at": imported_at,
    }


def _upsert_rows(connection, sheet_key, sheet_config, rows):
    schema = config.TRANSACTION_SCHEMA
    table = sheet_config["table"]
    column = sheet_config["column"]
    imported_at = datetime.now(timezone.utc).isoformat()

    inserted = 0
    updated = 0
    with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        for row in rows:
            meta = _metadata(sheet_config, row, imported_at)
            cursor.execute(
                f'''
                SELECT id
                FROM "{schema}"."{table}"
                WHERE lower(trim("{column}")) = lower(trim(%s))
                  AND COALESCE(__nc_deleted, false) = false
                LIMIT 1
                ''',
                (row["name"],),
            )
            existing = cursor.fetchone()
            if existing:
                cursor.execute(
                    f'''
                    UPDATE "{schema}"."{table}"
                    SET "{column}" = %s,
                        updated_at = now(),
                        nc_row_meta = %s::jsonb
                    WHERE id = %s
                    ''',
                    (row["name"], psycopg2.extras.Json(meta), existing["id"]),
                )
                updated += 1
            else:
                cursor.execute(
                    f'''
                    INSERT INTO "{schema}"."{table}"
                      (created_at, updated_at, "__nc_deleted", nc_row_meta, "{column}")
                    VALUES
                      (now(), now(), false, %s::jsonb, %s)
                    ''',
                    (psycopg2.extras.Json(meta), row["name"]),
                )
                inserted += 1
    connection.commit()
    return {"sheet": sheet_key, "read": len(rows), "inserted": inserted, "updated": updated}


def import_master_sheets(selected=None):
    selected = selected or tuple(MASTER_SHEETS)
    results = []
    with _connect() as connection:
        for sheet_key in selected:
            sheet_config = MASTER_SHEETS[sheet_key]
            rows = _read_single_column_workbook(sheet_config["path"], sheet_config["expected_header"])
            results.append(_upsert_rows(connection, sheet_key, sheet_config, rows))
    return results


def main():
    parser = argparse.ArgumentParser(description="Import customer/category master sheets into NocoDB tables.")
    parser.add_argument(
        "sheet",
        nargs="*",
        help="Optional sheet(s) to import. Defaults to customers and categories.",
    )
    args = parser.parse_args()
    invalid = sorted(set(args.sheet) - set(MASTER_SHEETS))
    if invalid:
        parser.error("invalid sheet(s): " + ", ".join(invalid))
    for result in import_master_sheets(tuple(args.sheet) if args.sheet else None):
        print(
            f"{result['sheet']}: read={result['read']} "
            f"inserted={result['inserted']} updated={result['updated']}"
        )


if __name__ == "__main__":
    main()
