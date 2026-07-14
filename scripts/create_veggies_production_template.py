#!/usr/bin/env python3
"""Generate the staff-facing wide Veggies Production Excel template."""

from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from tools.veggies_production import CropDefinition, load_crop_definitions


DEFAULT_OUTPUT = PROJECT_ROOT / "excel_import" / "BigShot_Veggies_Production_Template.xlsx"
ENTRY_ROWS = 500
GREEN = "174F3B"
LIGHT_GREEN = "E4F0EA"
GOLD = "D6A84B"
WHITE = "FFFFFF"
LIGHT_GREY = "E8ECEA"


def _style_header(sheet, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = sheet.cell(1, column)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34


def generate_template(output_path: str | Path = DEFAULT_OUTPUT,
                      crops: list[CropDefinition] | None = None) -> Path:
    """Create the template using active database crops, or supplied crops in tests."""
    crop_rows = crops if crops is not None else load_crop_definitions()
    crop_rows = list(crop_rows)
    if not crop_rows:
        raise ValueError("At least one active crop is required to generate the template.")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    entry = workbook.create_sheet("Data Entry")
    crop_sheet = workbook.create_sheet("Crop Master")

    instructions.sheet_view.showGridLines = False
    instructions["A1"] = "BigShot Veggies Production Import"
    instructions["A1"].font = Font(size=20, bold=True, color=GREEN)
    instructions["A3"] = "Purpose"
    instructions["A3"].font = Font(size=13, bold=True, color=GREEN)
    instructions["A4"] = "Record one daily vegetable production submission per row."
    instructions["A6"] = "Required fields"
    instructions["A6"].font = Font(size=13, bold=True, color=GREEN)
    instructions["A7"] = "Production Date and at least one crop quantity."
    instructions["A9"] = "Entry rules"
    instructions["A9"].font = Font(size=13, bold=True, color=GREEN)
    rules = [
        "Use YYYY/MM/DD for Production Date and Date of Entry.",
        "Crop quantities may be whole numbers or decimals and cannot be negative.",
        "Blank means not entered/no production record; 0 means confirmed zero production.",
        "Do not rename, add, or delete columns in the Data Entry table.",
        "Assignee, Note, AI Note, and Date of Entry are optional.",
        "Preview/Dry Run the workbook before confirming an import.",
    ]
    for row, text in enumerate(rules, start=10):
        instructions.cell(row, 1, f"• {text}")
    instructions.column_dimensions["A"].width = 105
    instructions.freeze_panes = "A3"
    instructions.protection.sheet = True

    headers = ["Production Date", *[crop.crop_name for crop in crop_rows],
               "Assignee", "Note", "AI Note", "Date of Entry"]
    entry.append(headers)
    examples = [
        [date(2026, 1, 1), 129, 48, 0.5],
        [date(2026, 1, 2), 0, None, 0.75],
        [date(2026, 1, 3), None, 32, None],
    ]
    for example_index, values in enumerate(examples, start=2):
        row = [None] * len(headers)
        row[0] = values[0]
        for offset, value in enumerate(values[1:], start=1):
            if offset < len(headers) - 4:
                row[offset] = value
        row[-4] = f"Example {example_index - 1}"
        row[-3] = "Replace or delete this example row"
        row[-1] = values[0]
        entry.append(row)
    for _ in range(4, ENTRY_ROWS + 2):
        entry.append([None] * len(headers))

    _style_header(entry, len(headers))
    entry["A1"].fill = PatternFill("solid", fgColor=GOLD)
    entry["A1"].font = Font(bold=True, color="17211C")
    entry.freeze_panes = "B2"
    entry.auto_filter.ref = f"A1:{entry.cell(ENTRY_ROWS + 1, len(headers)).coordinate}"
    entry.column_dimensions["A"].width = 18
    for column in range(2, 2 + len(crop_rows)):
        entry.column_dimensions[entry.cell(1, column).column_letter].width = 17
    for column in range(len(headers) - 3, len(headers) + 1):
        entry.column_dimensions[entry.cell(1, column).column_letter].width = 24
    for row in range(2, ENTRY_ROWS + 2):
        entry.cell(row, 1).number_format = "yyyy/mm/dd"
        entry.cell(row, len(headers)).number_format = "yyyy/mm/dd"
    quantity_range = entry.cell(2, 2).coordinate + ":" + entry.cell(ENTRY_ROWS + 1, 1 + len(crop_rows)).coordinate
    quantity_validation = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    quantity_validation.error = "Enter a number greater than or equal to zero, or leave the cell blank."
    quantity_validation.errorTitle = "Invalid quantity"
    quantity_validation.prompt = "Blank = not entered; 0 = confirmed zero production."
    quantity_validation.promptTitle = "Crop quantity"
    quantity_validation.showErrorMessage = True
    quantity_validation.showInputMessage = True
    entry.add_data_validation(quantity_validation)
    quantity_validation.add(quantity_range)
    entry.conditional_formatting.add(
        quantity_range,
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="F4CCCC")),
    )
    entry.cell(1, 1).border = Border(bottom=Side(style="medium", color=GOLD))
    for cell in entry[1]:
        cell.border = Border(bottom=Side(style="medium", color=GOLD))
    table = Table(displayName="VeggiesProductionEntry", ref=f"A1:{entry.cell(ENTRY_ROWS + 1, len(headers)).coordinate}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True, showColumnStripes=False)
    entry.add_table(table)

    crop_sheet.append(["Crop Code", "Crop Name", "Default Unit", "Active", "Display Order", "Accepted Headers"])
    for display_order, crop in enumerate(crop_rows, start=1):
        crop_sheet.append([
            crop.crop_code,
            crop.crop_name,
            crop.default_unit,
            "Yes",
            display_order,
            ", ".join(dict.fromkeys([crop.crop_name, crop.source_header])),
        ])
    _style_header(crop_sheet, 6)
    crop_sheet.freeze_panes = "A2"
    crop_sheet.column_dimensions["A"].width = 24
    crop_sheet.column_dimensions["B"].width = 28
    crop_sheet.column_dimensions["C"].width = 18
    crop_sheet.column_dimensions["D"].width = 12
    crop_sheet.column_dimensions["E"].width = 16
    crop_sheet.column_dimensions["F"].width = 45
    crop_sheet.protection.sheet = True
    for row in crop_sheet.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

    workbook.save(output_path)
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    print(generate_template(args.output))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
