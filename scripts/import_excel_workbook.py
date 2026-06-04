import argparse
import os
import sys
from datetime import date, datetime
from pathlib import Path

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from openpyxl import load_workbook

from tools.excel_importer import TABLES, import_excel_payload


SHEETS = {
    "transection": "Transection",
    "sotephwar_transection": "Sotephwar_Transection",
    "financial_obligations": "Financial_Obligations",
    "sotephwar_inventory": "Sotephwar_Inventory",
}

STATUS_COLUMNS = ["Upload_Status", "Uploaded_ID", "Uploaded_At", "Upload_Error"]


def _header_map(sheet):
    return {
        str(cell.value).strip(): index
        for index, cell in enumerate(sheet[1], start=1)
        if cell.value
    }


def _row_has_data(sheet, row_number, headers):
    for column in TABLES[_table_key_for_sheet(sheet.title)]["columns"]:
        index = headers.get(column)
        if index and sheet.cell(row_number, index).value not in (None, ""):
            return True
    return False


def _table_key_for_sheet(sheet_name):
    for table_key, candidate in SHEETS.items():
        if candidate == sheet_name:
            return table_key
    raise KeyError(sheet_name)


def _read_payload(workbook):
    payload = {}
    for table_key, sheet_name in SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            payload[table_key] = []
            continue

        sheet = workbook[sheet_name]
        headers = _header_map(sheet)
        rows = []
        status_column = headers.get("Upload_Status")

        for row_number in range(2, sheet.max_row + 1):
            status = sheet.cell(row_number, status_column).value if status_column else None
            if str(status or "").strip().upper() == "INSERTED":
                continue
            if not _row_has_data(sheet, row_number, headers):
                continue

            row = {"__row_number": row_number}
            for column in TABLES[table_key]["columns"]:
                index = headers.get(column)
                cell = sheet.cell(row_number, index) if index else None
                row[column] = _cell_value(table_key, column, cell)
            rows.append(row)

        payload[table_key] = rows
    return payload


def _cell_value(table_key, column, cell):
    if cell is None:
        return None

    value = cell.value
    if column not in TABLES[table_key]["date"]:
        return value

    if isinstance(value, datetime):
        value = value.date()

    if (
        isinstance(value, date)
        and _uses_year_day_month_format(cell.number_format)
        and value.day <= 12
        and value.month <= 12
    ):
        return date(value.year, value.day, value.month)

    return value


def _uses_year_day_month_format(number_format):
    fmt = str(number_format or "").lower()
    fmt = fmt.replace("\\", "").replace("-", "/").replace(".", "/")
    fmt = fmt.replace(" ", "")
    return "yyyy/dd/mm" in fmt or "yy/dd/mm" in fmt


def _ensure_status_columns(sheet):
    headers = _header_map(sheet)
    next_column = sheet.max_column + 1
    for column in STATUS_COLUMNS:
        if column not in headers:
            sheet.cell(1, next_column).value = column
            headers[column] = next_column
            next_column += 1
    return headers


def _write_results(workbook, results):
    uploaded_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for table_key, result in results.items():
        sheet_name = SHEETS[table_key]
        if sheet_name not in workbook.sheetnames:
            continue

        sheet = workbook[sheet_name]
        headers = _ensure_status_columns(sheet)

        for row in result["inserted"]:
            row_number = row.get("row_number")
            if not row_number:
                continue
            sheet.cell(row_number, headers["Upload_Status"]).value = "INSERTED"
            sheet.cell(row_number, headers["Uploaded_ID"]).value = row.get("id")
            sheet.cell(row_number, headers["Uploaded_At"]).value = uploaded_at
            sheet.cell(row_number, headers["Upload_Error"]).value = ""

        for row in result["errors"]:
            row_number = row.get("row_number")
            if not row_number:
                continue
            sheet.cell(row_number, headers["Upload_Status"]).value = "ERROR"
            sheet.cell(row_number, headers["Uploaded_At"]).value = uploaded_at
            sheet.cell(row_number, headers["Upload_Error"]).value = row.get("error")


def _print_summary(payload, results=None):
    for table_key, sheet_name in SHEETS.items():
        rows = payload.get(table_key, [])
        if results is None:
            print(f"{sheet_name}: {len(rows)} row(s) ready")
            continue
        result = results[table_key]
        print(
            f"{sheet_name}: {len(result['inserted'])} inserted, "
            f"{len(result['errors'])} error(s)"
        )
        for error in result["errors"]:
            print(f"  row {error.get('row_number')}: {error.get('error')}")


def main():
    parser = argparse.ArgumentParser(description="Append Excel rows into business tables.")
    parser.add_argument(
        "workbook",
        nargs="?",
        default=str(Path(PROJECT_ROOT) / "excel_import" / "business_data_import_template.xlsx"),
    )
    parser.add_argument("--dry-run", action="store_true", help="Count rows without inserting.")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser()
    workbook = load_workbook(workbook_path)
    payload = _read_payload(workbook)

    if args.dry_run:
        _print_summary(payload)
        return 0

    results = import_excel_payload(payload)
    _write_results(workbook, results)
    workbook.save(workbook_path)
    _print_summary(payload, results)

    if any(result["errors"] for result in results.values()):
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
