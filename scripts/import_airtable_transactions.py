import argparse
import csv
import os
import re
import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(PROJECT_ROOT) / "excel_import" / "business_data_import_template.xlsx"
DEFAULT_INPUTS = [
    ("Farm", Path.home() / "Desktop" / "June farm.csv"),
    ("Sote Phwar", Path.home() / "Desktop" / "June sotephwar.csv"),
]
TRANSECTION_COLUMNS = [
    "Date",
    "Income_Expense",
    "Categorization",
    "Sector",
    "Item_Description",
    "Amount",
    "Payment_Method",
    "Attachement",
    "AI_Comment",
]
STATUS_COLUMNS = ["Upload_Status", "Uploaded_ID", "Uploaded_At", "Upload_Error"]


def _header_map(sheet):
    return {
        str(cell.value).strip(): index
        for index, cell in enumerate(sheet[1], start=1)
        if cell.value
    }


def _clean_text(value):
    value = str(value or "").replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", value)


def _parse_date(value):
    value = _clean_text(value)
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unsupported date: {value}")


def _parse_amount(value):
    value = _clean_text(value)
    if not value:
        return None
    value = value.replace("K", "").replace(",", "")
    return int(Decimal(value))


def _amount_for_row(row):
    income_expense = _clean_text(row.get("Income/Expense"))
    if income_expense.lower() == "income":
        return (
            _parse_amount(row.get("Income Amount (Kyats)"))
            or _parse_amount(row.get("Total Income"))
        )
    return (
        _parse_amount(row.get("Expend Amount (Kyats)"))
        or _parse_amount(row.get("Total Expense"))
    )


def _note_for_row(row):
    parts = [
        _clean_text(row.get("Notes")),
        _clean_text(row.get("Plan")),
    ]
    return " | ".join(part for part in parts if part) or None


def _csv_row_to_transection(row, sector):
    return {
        "Date": _parse_date(row.get("Date")),
        "Income_Expense": _clean_text(row.get("Income/Expense")),
        "Categorization": _clean_text(row.get("Categorization")),
        "Sector": sector,
        "Item_Description": _clean_text(row.get("Items")),
        "Amount": _amount_for_row(row),
        "Payment_Method": _clean_text(row.get("Method of purchase")),
        "Attachement": _clean_text(row.get("Photo (optional)")) or None,
        "AI_Comment": _note_for_row(row),
    }


def _key(row):
    return tuple(_normal_key_value(row.get(column)) for column in TRANSECTION_COLUMNS[:7])


def _normal_key_value(value):
    if hasattr(value, "date"):
        value = value.date()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, (int, float, Decimal)):
        return str(int(value))
    return _clean_text(value).lower()


def _existing_keys(sheet, headers):
    keys = set()
    for row_number in range(2, sheet.max_row + 1):
        row = {}
        has_data = False
        for column in TRANSECTION_COLUMNS:
            index = headers.get(column)
            value = sheet.cell(row_number, index).value if index else None
            row[column] = value
            if value not in (None, ""):
                has_data = True
        if has_data:
            keys.add(_key(row))
    return keys


def _next_append_row(sheet, headers):
    date_column = headers["Date"]
    for row_number in range(2, sheet.max_row + 1):
        if sheet.cell(row_number, date_column).value in (None, ""):
            return row_number
    return sheet.max_row + 1


def _append_row(sheet, headers, row, row_number):
    for column in TRANSECTION_COLUMNS:
        sheet.cell(row_number, headers[column]).value = row.get(column)
    for column in STATUS_COLUMNS:
        index = headers.get(column)
        if index:
            sheet.cell(row_number, index).value = None

    sheet.cell(row_number, headers["Date"]).number_format = "dd/mm/yyyy"


def _load_csv_rows(path, sector):
    with path.open(newline="", encoding="utf-8-sig") as csv_file:
        return [_csv_row_to_transection(row, sector) for row in csv.DictReader(csv_file)]


def _input_arg(value):
    if "=" not in value:
        raise argparse.ArgumentTypeError("Use Sector=/path/to/file.csv")
    sector, path = value.split("=", 1)
    sector = _clean_text(sector)
    if not sector:
        raise argparse.ArgumentTypeError("Sector cannot be blank")
    return sector, Path(path).expanduser()


def main():
    parser = argparse.ArgumentParser(
        description="Convert Airtable transaction CSV exports into the Excel import workbook."
    )
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument(
        "--input",
        action="append",
        type=_input_arg,
        help="CSV input as Sector=/path/to/file.csv. Defaults to Desktop June farm/sotephwar.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser()
    inputs = args.input or DEFAULT_INPUTS

    workbook = load_workbook(workbook_path)
    sheet = workbook["Transection"]
    headers = _header_map(sheet)
    missing = [column for column in TRANSECTION_COLUMNS if column not in headers]
    if missing:
        raise ValueError("Workbook Transection sheet is missing: " + ", ".join(missing))

    known_keys = _existing_keys(sheet, headers)
    append_row = _next_append_row(sheet, headers)
    summary = []

    for sector, csv_path in inputs:
        rows = _load_csv_rows(csv_path, sector)
        added = 0
        skipped = 0
        for row in rows:
            row_key = _key(row)
            if row_key in known_keys:
                skipped += 1
                continue
            _append_row(sheet, headers, row, append_row)
            known_keys.add(row_key)
            append_row += 1
            added += 1
        summary.append((sector, csv_path, added, skipped))

    workbook.save(workbook_path)

    for sector, csv_path, added, skipped in summary:
        print(f"{sector}: {added} added, {skipped} skipped from {csv_path}")
    print(f"Workbook: {workbook_path}")


if __name__ == "__main__":
    main()
