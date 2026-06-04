import os
import sys
from pathlib import Path

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from tools.excel_importer import TABLES


OUTPUT_DIR = Path(PROJECT_ROOT) / "excel_import"
WORKBOOK_PATH = OUTPUT_DIR / "business_data_import_template.xlsx"


SHEETS = [
    ("transection", "Transection"),
    ("sotephwar_transection", "Sotephwar_Transection"),
    ("financial_obligations", "Financial_Obligations"),
    ("sotephwar_inventory", "Sotephwar_Inventory"),
]


def _write_sheet(workbook, table_key, sheet_name):
    sheet = workbook.create_sheet(sheet_name)
    columns = TABLES[table_key]["columns"]
    headers = columns + ["Upload_Status", "Uploaded_ID", "Uploaded_At", "Upload_Error"]
    sheet.append(headers)

    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for _ in range(100):
        sheet.append([""] * len(headers))

    date_columns = set(TABLES[table_key]["date"])
    for column_index, header in enumerate(headers, start=1):
        if header in date_columns:
            for row_index in range(2, 102):
                sheet.cell(row_index, column_index).number_format = "dd/mm/yyyy"

    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 14), 34)

    sheet.freeze_panes = "A2"
    table_ref = f"A1:{sheet.cell(row=101, column=len(headers)).coordinate}"
    excel_table = Table(displayName=sheet_name.replace("_", "") + "Import", ref=table_ref)
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(excel_table)


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    instructions = workbook.create_sheet("Instructions")
    instructions["A1"] = "Business Data Import"
    instructions["A1"].font = Font(bold=True, size=16)
    instructions["A3"] = "Fill rows in the three table sheets. Leave Upload_Status blank for rows you want to insert."
    instructions["A4"] = "Run scripts/excel_import_server.py, then click the Excel macro button or run UploadBusinessData."
    instructions["A5"] = "Rows marked INSERTED are skipped by the macro, so clicking again does not resend those rows."
    instructions["A6"] = "The importer only inserts new rows. It does not delete, update, truncate, or replace NocoDB/Postgres data."
    instructions.column_dimensions["A"].width = 120

    for table_key, sheet_name in SHEETS:
        _write_sheet(workbook, table_key, sheet_name)

    workbook.save(WORKBOOK_PATH)
    print(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
