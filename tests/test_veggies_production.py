import io
import json
import tempfile
import unittest
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from tools.veggies_production import (
    CropDefinition,
    crop_header_map,
    default_crop_definitions,
    import_veggies_preview,
    normalize_header,
    parse_production_date,
    parse_quantity,
    parse_veggies_workbook,
)
from scripts.create_veggies_production_template import generate_template


def workbook_bytes(headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data Entry"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


class ExistingImportCursor:
    def __init__(self):
        self.last_sql = ""

    def __enter__(self):
        return self

    def __exit__(self, *args):
        return False

    def execute(self, sql, params=None):
        self.last_sql = sql

    def fetchone(self):
        return {"id": 42}


class ExistingImportConnection:
    def __init__(self):
        self.rolled_back = False

    def cursor(self, **kwargs):
        return ExistingImportCursor()

    def rollback(self):
        self.rolled_back = True


class FailingCursor:
    def __enter__(self):
        return self

    def __exit__(self, *args):
        return False

    def execute(self, sql, params=None):
        raise RuntimeError("database failure")


class FailingConnection:
    def __init__(self):
        self.rolled_back = False

    def cursor(self, **kwargs):
        return FailingCursor()

    def rollback(self):
        self.rolled_back = True


class VeggiesProductionTest(unittest.TestCase):
    def test_excel_serial_date_conversion(self):
        self.assertEqual(date(2026, 1, 1), parse_production_date(46023))
        self.assertEqual(date(2026, 1, 2), parse_production_date(46024))

    def test_yyyy_slash_date_conversion(self):
        self.assertEqual(date(2026, 7, 14), parse_production_date("2026/07/14"))

    def test_dd_mm_yyyy_date_conversion(self):
        self.assertEqual(date(2026, 7, 14), parse_production_date("14/07/2026"))

    def test_true_excel_date_cell_value(self):
        self.assertEqual(date(2026, 7, 14), parse_production_date(datetime(2026, 7, 14)))

    def test_header_alias_normalization(self):
        crop_map = crop_header_map(default_crop_definitions())
        self.assertEqual("Parsley", crop_map[normalize_header("Persley")].crop_name)
        self.assertEqual("Swiss Chard", crop_map[normalize_header("Swiss Chert")].crop_name)
        self.assertEqual("Fennel Bulb", crop_map[normalize_header("Funnel Bulb")].crop_name)

    def test_blank_quantity_creates_no_item_but_zero_does(self):
        preview = parse_veggies_workbook(workbook_bytes(
            ["Date", "Zucchini", "Cherry Tomato"],
            [["2026/01/01", None, 0]],
        ))
        self.assertEqual(1, preview.accepted_rows)
        self.assertEqual(1, preview.item_count)
        self.assertEqual(Decimal("0"), preview.valid_rows[0].items[0]["quantity"])

    def test_decimal_quantity_is_preserved(self):
        self.assertEqual(Decimal("0.5"), parse_quantity(0.5, "Rosemary"))

    def test_negative_quantity_is_rejected(self):
        with self.assertRaisesRegex(ValueError, "cannot be negative"):
            parse_quantity(-1, "Zucchini")

    def test_invalid_crop_value_is_rejected_with_row_and_column(self):
        preview = parse_veggies_workbook(workbook_bytes(
            ["Production Date", "Cherry Tomato"],
            [["2026-01-01", "abc"]],
        ))
        self.assertEqual(0, preview.accepted_rows)
        self.assertEqual(1, preview.rejected_rows)
        self.assertEqual(2, preview.errors[0].row_number)
        self.assertEqual("Cherry Tomato", preview.errors[0].column)

    def test_unknown_crop_header_is_rejected(self):
        preview = parse_veggies_workbook(workbook_bytes(
            ["Date", "Broccoli"],
            [["2026-01-01", 2]],
        ))
        self.assertEqual(["Broccoli"], preview.unknown_headers)
        self.assertEqual(0, preview.accepted_rows)

    def test_duplicate_source_row_is_rejected_but_same_date_is_allowed(self):
        preview = parse_veggies_workbook(workbook_bytes(
            ["Date", "Zucchini", "Carrot"],
            [
                ["2026-01-01", 2, None],
                ["2026-01-01", 2, None],
                ["2026-01-01", None, 2],
            ],
        ))
        self.assertEqual(2, preview.accepted_rows)
        self.assertEqual([3], preview.duplicate_rows)

    def test_wide_row_normalizes_to_one_batch_and_multiple_items(self):
        preview = parse_veggies_workbook(workbook_bytes(
            ["Date", "Zucchini", "Rosemary", "Asignee", "Note"],
            [["2026-01-01", 129, 0.5, "Aye Aye", "morning harvest"]],
        ))
        row = preview.valid_rows[0]
        self.assertEqual(date(2026, 1, 1), row.production_date)
        self.assertEqual("Aye Aye", row.assignee)
        self.assertEqual(2, len(row.items))

    def test_second_identical_file_import_reports_already_exists(self):
        preview = parse_veggies_workbook(workbook_bytes(
            ["Date", "Zucchini"], [["2026-01-01", 1]],
        ))
        connection = ExistingImportConnection()
        result = import_veggies_preview(preview, connection=connection)
        self.assertTrue(result["already_exists"])
        self.assertEqual(42, result["import_id"])
        self.assertTrue(connection.rolled_back)

    def test_fatal_database_failure_rolls_back(self):
        preview = parse_veggies_workbook(workbook_bytes(
            ["Date", "Zucchini"], [["2026-01-01", 1]],
        ))
        connection = FailingConnection()
        with self.assertRaisesRegex(RuntimeError, "database failure"):
            import_veggies_preview(preview, connection=connection)
        self.assertTrue(connection.rolled_back)

    def test_migration_is_additive_and_never_mentions_legacy_farm_object(self):
        migration = (
            Path(__file__).resolve().parents[1]
            / "migrations"
            / "20260714_001_veggies_production_up.sql"
        ).read_text()
        self.assertIn("veggies_production_batches", migration)
        self.assertIn("veggies_production_items", migration)
        self.assertNotIn("farm_transection", migration)

    def test_template_generation_has_required_structure_and_validation(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "BigShot_Veggies_Production_Template.xlsx"
            generate_template(output, default_crop_definitions())
            from openpyxl import load_workbook
            workbook = load_workbook(output)
            self.assertEqual(["Instructions", "Data Entry", "Crop Master"], workbook.sheetnames)
            entry = workbook["Data Entry"]
            headers = [cell.value for cell in entry[1]]
            self.assertEqual("Production Date", headers[0])
            self.assertIn("Iceberg Lettuce", headers)
            self.assertNotIn("Farm Area", headers)
            self.assertEqual("Date of Entry", headers[-1])
            self.assertTrue(entry.data_validations.count)
            self.assertEqual("yyyy/mm/dd", entry["A2"].number_format)
            self.assertTrue(workbook["Instructions"].protection.sheet)
            self.assertTrue(workbook["Crop Master"].protection.sheet)

    def test_nocodb_metadata_exposes_only_two_normal_user_tables(self):
        metadata_path = Path(__file__).resolve().parents[1] / "nocodb" / "veggies_production_metadata.json"
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        visible = [table["title"] for table in metadata["visible_tables"]]
        self.assertEqual(["Veggies Production Entry", "Veggies Crop Master"], visible)
        self.assertIn("veggies_production_items", metadata["hidden_tables"])
        self.assertIn("farm_production", metadata["legacy_objects_must_remain_unchanged"])


if __name__ == "__main__":
    unittest.main()
