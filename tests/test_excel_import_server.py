import io
import unittest
from unittest.mock import patch

from openpyxl import Workbook

from scripts import excel_import_server


def upload_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Date", "Zucchini"])
    sheet.append(["2026/01/01", 5])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


class ExcelImportServerTest(unittest.TestCase):
    def setUp(self):
        excel_import_server._VEGGIES_PREVIEWS.clear()
        excel_import_server._AUDIT_IMPORT_SESSIONS.clear()
        self.client = excel_import_server.app.test_client()

    def test_import_page_directs_veggies_to_browser_portal(self):
        response = self.client.get("/")
        self.assertEqual(200, response.status_code)
        self.assertIn(b"Veggies Production Basic", response.data)
        self.assertIn(b'href="/business-os/veggies-production"', response.data)
        self.assertNotIn(b"127.0.0.1:5059", response.data)
        self.assertIn(b"Optional legacy utilities", response.data)
        self.assertIn(b"Audit Existing Records", response.data)
        self.assertNotIn(b"Import New Records", response.data)
        self.assertNotIn(b"Reconcile and Update", response.data)

    @patch("scripts.excel_import_server.data_audit.upload_audit")
    def test_unified_data_audit_session_is_preview_only(self, upload_audit):
        upload_audit.return_value = {
            "id": 41, "filename": "source.xlsx", "sheet_name": "Sheet1",
            "source_rows": 8, "source_columns": 4, "detected_date": "2026-01-01",
            "status": "ready",
        }
        response = self.client.post(
            "/data-audit/session",
            data={
                "operation": "audit",
                "target_key": "sotephwar_transection",
                "file": (io.BytesIO(b"Date,Amount\n2026-01-01,1\n"), "source.csv"),
            }, content_type="multipart/form-data",
        )
        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertEqual("audit", payload["operation"])
        self.assertEqual("SotePhwar Transaction", payload["target_label"])
        upload_audit.assert_called_once()

    @patch("scripts.excel_import_server.import_excel_payload")
    def test_existing_json_import_endpoint_remains_functional(self, import_payload):
        import_payload.return_value = {
            "transection": {"inserted": [{"id": 1}], "errors": []},
        }
        response = self.client.post("/import", json={"transection": [{"Amount": 1}]})
        self.assertEqual(200, response.status_code)
        self.assertTrue(response.get_json()["ok"])

    @patch("scripts.excel_import_server.load_crop_definitions")
    def test_veggies_preview_endpoint_returns_normalized_counts(self, crop_definitions):
        from tools.veggies_production import default_crop_definitions
        crop_definitions.return_value = default_crop_definitions()
        response = self.client.post(
            "/veggies-production/preview",
            data={"workbook": (upload_workbook(), "production.xlsx")},
            content_type="multipart/form-data",
        )
        payload = response.get_json()
        self.assertEqual(200, response.status_code)
        self.assertEqual(1, payload["accepted_rows"])
        self.assertEqual(1, payload["normalized_items"])
        self.assertIn(payload["preview_id"], excel_import_server._VEGGIES_PREVIEWS)


if __name__ == "__main__":
    unittest.main()
