import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from tools.bi_reports import write_excel_report


class BiReportsTest(unittest.TestCase):
    def test_kpi_excel_serializes_nested_sources(self):
        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "kpi.xlsx"
            write_excel_report(
                {
                    "title": "BigShot KPI Report",
                    "period_label": "This Month",
                    "result": {
                        "formula": "kpi_overview",
                        "total_income": 1000,
                        "sources": {
                            "farm_transection_total_amount": 400,
                            "sotephwar_transection_total_amount": 600,
                        },
                    },
                },
                output_path,
            )

            workbook = load_workbook(output_path, read_only=True)
            rows = list(workbook["Report"].iter_rows(values_only=True))

        source_row = next(row for row in rows if row[0] == "sources")
        self.assertIn('"farm_transection_total_amount": 400', source_row[1])
        self.assertIn('"sotephwar_transection_total_amount": 600', source_row[1])


if __name__ == "__main__":
    unittest.main()
