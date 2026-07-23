"""Permanent, metadata-driven Data Audit Center service."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import psycopg2.extras
from psycopg2 import sql
from openpyxl import load_workbook
from openpyxl.utils.datetime import CALENDAR_WINDOWS_1900, from_excel

import config
from tools import excel_importer, formula_engine


SCHEMA = config.TRANSACTION_SCHEMA
AUDIT_TABLE = "business_os_data_audit"
ROW_TABLE = "business_os_data_audit_row"
ALIAS_TABLE = "business_os_data_audit_alias"
MAPPING_TABLE = "business_os_data_audit_mapping"
BACKUP_TABLE = "business_os_data_audit_backup"

COMMON_ALIASES = {
    "invoice_number": {"invoice number", "invoice no", "voucher number", "voucher no"},
    "invoice_date": {"invoice date", "voucher date", "date"},
    "customer_name": {"customer name", "customer"},
    "product": {"item", "product", "bottle type", "item description", "description"},
    "quantity": {"quantity", "qty"},
    "total_amount": {"total amount", "invoice amount", "amount"},
    "amount_received": {"amount received", "received", "paid", "total received"},
    "category": {"category", "categorization"},
    "transaction_type": {"income expense", "income or expense", "type"},
    "sector": {"sector", "business unit"},
    "payment_method": {"payment method", "method"},
    "note": {"note", "notes", "remark", "remarks"},
    "ai_comment": {"ai comment", "ai comments", "ai analysis"},
}

TARGETS = {
    "sotephwar_transection": {
        "label": "SotePhwar Transaction",
        "table": "Sotephwar_Transection",
        "import_key": "sotephwar_transection",
        "business_date_field": "Invoice_Date",
        "amount_interpretation": "sum_line_amounts_once",
        "ignored_metadata_fields": {"created_at", "inserted_at", "updated_at", "import_date", "submission_date", "record_creation_timestamp"},
        "fields": {
            "invoice_number": "Invoice_Number", "invoice_date": "Invoice_Date",
            "customer_name": "Customer_Name", "product": "Item", "quantity": "Quantity",
            "total_amount": "Total_Amount", "amount_received": "Total_Received", "note": "Note", "ai_comment": "AI_comment",
        },
        "required": {"invoice_number", "invoice_date", "customer_name", "product", "quantity", "total_amount"},
        "match": ("invoice_number", "invoice_date", "customer_name", "product"),
        "compare": ("quantity", "total_amount"),
        "alias_fields": {"customer_name": "customer", "product": "product"},
        "update_fields": {"invoice_number", "invoice_date", "customer_name", "product", "quantity", "total_amount", "note", "ai_comment"},
    },
    "farm_transection": {
        "label": "Farm Transaction",
        "table": "farm_transection",
        "import_key": "farm_transection",
        "business_date_field": "Date",
        "amount_interpretation": "sum_line_amounts_once",
        "ignored_metadata_fields": {"created_at", "inserted_at", "updated_at", "import_date", "submission_date", "record_creation_timestamp"},
        "fields": {
            "invoice_number": "Invoice_Number", "invoice_date": "Date",
            "customer_name": "Customer", "total_amount": "Total_Amount",
            "amount_received": "Total_Received", "note": "Note", "ai_comment": "AI_Analysis",
        },
        "required": {"invoice_number", "invoice_date", "customer_name", "total_amount"},
        "match": ("invoice_number", "invoice_date", "customer_name"),
        "compare": ("total_amount",),
        "alias_fields": {"customer_name": "customer"},
        "update_fields": {"invoice_number", "invoice_date", "customer_name", "total_amount", "note", "ai_comment"},
    },
    "transection": {
        "label": "General Transaction",
        "table": "Transection",
        "import_key": "transection",
        "business_date_field": "Date",
        "amount_interpretation": "one_amount_per_transaction_row",
        "ignored_metadata_fields": {"created_at", "inserted_at", "updated_at", "import_date", "submission_date", "record_creation_timestamp"},
        "fields": {
            "invoice_date": "Date", "transaction_type": "Income_Expense", "category": "Categorization",
            "sector": "Sector", "product": "Item_Description", "total_amount": "Amount",
            "payment_method": "Payment_Method", "ai_comment": "AI_Comment",
        },
        "required": {"invoice_date", "transaction_type", "category", "product", "total_amount"},
        "match": ("invoice_date", "transaction_type", "category", "sector", "product"),
        "compare": ("total_amount",),
        "alias_fields": {"product": "product"},
        "update_fields": {"invoice_date", "transaction_type", "category", "sector", "product", "total_amount", "payment_method", "ai_comment"},
    },
}


def _connect():
    return formula_engine._connect()


def _table(name):
    return sql.SQL("{}.{}").format(sql.Identifier(SCHEMA), sql.Identifier(name))


def _text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value).strip()


def _json_value(value):
    if isinstance(value, dict):
        return {str(key): _json_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_value(item) for item in value]
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return _text(value)
    return value


def _json(value):
    return json.dumps(_json_value(value), ensure_ascii=False)


def normalize_header(value):
    value = re.sub(r"([a-z])([A-Z])", r"\1 \2", _text(value))
    value = value.replace("_", " ").replace("-", " ")
    return re.sub(r"\s+", " ", value).strip().casefold()


def normalize_lookup(value):
    value = re.sub(r"\s*,\s*", ",", _text(value))
    return re.sub(r"\s+", " ", value).strip().casefold()


def _normalize_invoice(value):
    value = _text(value)
    return re.sub(r"\.0+$", "", value).casefold()


def _normalize_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    value = _text(value)
    if not value:
        return ""
    # Excel uploads may persist datetimes as ISO strings.  Strip the timezone
    # marker and time portion before applying the business-date comparison.
    iso_value = value.strip().replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(iso_value)
        return parsed.date().isoformat()
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    # CSV exports sometimes contain the Excel serial number rather than a
    # formatted date.  Only accept a plausible date serial range.
    try:
        serial = float(value)
        if 1 <= serial <= 2958465:
            parsed = from_excel(serial, epoch=CALENDAR_WINDOWS_1900)
            if isinstance(parsed, datetime):
                parsed = parsed.date()
            return parsed.isoformat()
    except (TypeError, ValueError, OverflowError):
        pass
    return value.casefold()


def _number(value):
    if value is None or _text(value) == "":
        return None
    try:
        return Decimal(_text(value).replace(",", ""))
    except InvalidOperation:
        return None


def _default_customer(value):
    original = re.sub(r"\s+", " ", _text(value)).strip(" ,")
    match = re.fullmatch(r"(.+?),\s*(Daw|Ma|U|Ko)", original, re.I)
    if match:
        title = {"daw": "Daw", "ma": "Ma", "u": "U", "ko": "Ko"}[match.group(2).casefold()]
        return f"{title} {match.group(1).strip()}"
    return re.sub(r"\s*,\s*", ", ", original)


def _default_product(value):
    original = re.sub(r"\s+", " ", _text(value)).strip()
    token = re.sub(r"[\s_-]+", "", original).casefold().replace("sotephwar", "")
    return {
        "1l": "Sote Phwar 1L", "4l": "Sote Phwar 4L",
        "500ml": "Sote Phwar 500 mL", "100ml": "Sote Phwar 100 mL",
    }.get(token, original)


def _load_aliases(connection, target_key):
    with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        cursor.execute(
            sql.SQL("SELECT alias_type,normalized_lookup,normalized_value FROM {} "
                    "WHERE active=true AND (target_key=%s OR target_key IS NULL) "
                    "ORDER BY target_key NULLS FIRST,id").format(_table(ALIAS_TABLE)),
            (target_key,),
        )
        return {
            (row["alias_type"], row["normalized_lookup"]): row["normalized_value"]
            for row in cursor.fetchall()
        }


def _normalize_field(field, value, aliases):
    if field == "invoice_number":
        return _normalize_invoice(value)
    if field == "invoice_date":
        return _normalize_date(value)
    if field in {"quantity", "total_amount", "amount_received"}:
        number = _number(value)
        return _text(number)
    if field == "customer_name":
        default = _default_customer(value)
        return aliases.get(("customer", normalize_lookup(value)), default).casefold()
    if field == "product":
        default = _default_product(value)
        return aliases.get(("product", normalize_lookup(value)), default).casefold()
    return normalize_lookup(value)


def detect_mapping(headers, target_key):
    target = TARGETS[target_key]
    result, confidence, conflicts = {}, {}, {}
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        candidates = [field for field in target["fields"] if normalized in COMMON_ALIASES.get(field, set())]
        if len(candidates) == 1:
            field = candidates[0]
            if field in result:
                conflicts[field] = [result[field], index]
                result.pop(field, None)
                confidence[field] = 0
            else:
                result[field] = index
                confidence[field] = 1.0
    missing = sorted(target["required"] - set(result))
    return result, confidence, missing, conflicts


def _read_upload(file_storage):
    filename = file_storage.filename or "upload"
    extension = filename.rsplit(".", 1)[-1].casefold() if "." in filename else ""
    raw = file_storage.stream.read()
    if extension == "csv":
        decoded = raw.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(decoded)))
        return "CSV", rows
    if extension == "xlsx":
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        return worksheet.title, [list(row) for row in worksheet.iter_rows(values_only=True)]
    if extension == "xls":
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError("Legacy .xls support requires the xlrd dependency.") from exc
        workbook = xlrd.open_workbook(file_contents=raw)
        worksheet = workbook.sheet_by_index(0)
        return worksheet.name, [worksheet.row_values(index) for index in range(worksheet.nrows)]
    raise ValueError("Only .xlsx, .xls and .csv files are supported.")


def upload_audit(file_storage, target_key, username):
    if target_key not in TARGETS:
        raise ValueError("Unsupported target table.")
    sheet_name, rows = _read_upload(file_storage)
    if not rows:
        raise ValueError("The uploaded file is empty.")
    headers = [_text(value) for value in rows[0]]
    if not any(headers):
        raise ValueError("The uploaded file has no headers.")
    mapping, confidence, missing, conflicts = detect_mapping(headers, target_key)
    source_data = [
        {"__row_number": index, **{headers[column]: _json_value(value) for column, value in enumerate(row) if column < len(headers)}}
        for index, row in enumerate(rows[1:], start=2)
        if any(_text(value) for value in row)
    ]
    dates = []
    date_index = mapping.get("invoice_date")
    if date_index is not None:
        for row in rows[1:]:
            if date_index < len(row):
                normalized = _normalize_date(row[date_index])
                if re.fullmatch(r"\d{4}-\d{2}-\d{2}", normalized):
                    dates.append(normalized)
    status = "mapping_required" if missing or conflicts else "ready"
    warnings = []
    if missing:
        warnings.append("Required mappings missing: " + ", ".join(missing))
    if conflicts:
        warnings.append("Duplicate logical mappings require review.")
    with _connect() as connection, connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        cursor.execute(
            sql.SQL("INSERT INTO {} (created_by,target_key,target_table,filename,sheet_name,source_rows,"
                    "source_columns,detected_date,headers,column_mapping,mapping_confidence,source_data,warnings,status) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb,%s::jsonb,%s::jsonb,%s::jsonb,%s::jsonb,%s) RETURNING id").format(_table(AUDIT_TABLE)),
            (username, target_key, TARGETS[target_key]["table"], file_storage.filename, sheet_name,
             len(source_data), len(headers), max(dates) if dates else None, _json(headers), _json(mapping),
             _json(confidence), _json(source_data), _json(warnings), status),
        )
        audit_id = cursor.fetchone()["id"]
        connection.commit()
    return get_audit(audit_id)


def get_audit(audit_id, connection=None):
    owns = connection is None
    connection = connection or _connect()
    try:
        with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(sql.SQL("SELECT * FROM {} WHERE id=%s").format(_table(AUDIT_TABLE)), (int(audit_id),))
            row = cursor.fetchone()
        if not row:
            raise LookupError("Audit not found.")
        value = dict(row)
        for key in ("created_at", "applied_at"):
            if value.get(key):
                value[key] = value[key].isoformat()
        return value
    finally:
        if owns:
            connection.close()


def save_mapping(audit_id, mapping, username, template_name=""):
    with _connect() as connection, connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        audit = get_audit(audit_id, connection)
        target = TARGETS[audit["target_key"]]
        clean = {field: int(index) for field, index in mapping.items() if field in target["fields"]}
        missing = sorted(target["required"] - set(clean))
        if missing:
            raise ValueError("Required mappings missing: " + ", ".join(missing))
        if len(set(clean.values())) != len(clean):
            raise ValueError("One source column cannot map to multiple fields.")
        cursor.execute(
            sql.SQL("UPDATE {} SET column_mapping=%s::jsonb,status='ready',version=version+1 WHERE id=%s RETURNING id").format(_table(AUDIT_TABLE)),
            (_json(clean), int(audit_id)),
        )
        if template_name.strip():
            normalized_headers = [normalize_header(value) for value in audit["headers"]]
            cursor.execute(
                sql.SQL("INSERT INTO {} (name,target_key,normalized_headers,column_mapping,created_by,updated_by) "
                        "VALUES (%s,%s,%s::jsonb,%s::jsonb,%s,%s) "
                        "ON CONFLICT (target_key,name) DO UPDATE SET normalized_headers=EXCLUDED.normalized_headers,"
                        "column_mapping=EXCLUDED.column_mapping,updated_at=now(),updated_by=EXCLUDED.updated_by").format(_table(MAPPING_TABLE)),
                (template_name.strip(), audit["target_key"], _json(normalized_headers), _json(clean), username, username),
            )
        connection.commit()
    return get_audit(audit_id)


def _database_rows(connection, target):
    columns = sorted(set(target["fields"].values()))
    select = sql.SQL(",").join(sql.Identifier(column) for column in columns)
    with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        cursor.execute(
            sql.SQL("SELECT id,{} FROM {} WHERE COALESCE(__nc_deleted,false)=false ORDER BY id").format(
                select, _table(target["table"])
            )
        )
        return [dict(row) for row in cursor.fetchall()]


def _logical_source_rows(audit, target):
    headers = audit["headers"]
    mapping = {key: int(value) for key, value in audit["column_mapping"].items()}
    result = []
    for source in audit["source_data"]:
        result.append({
            "__row_number": source["__row_number"],
            **{field: source.get(headers[index]) for field, index in mapping.items()},
        })
    return result


def run_audit(audit_id, username):
    with _connect() as connection:
        connection.set_session(readonly=False, isolation_level="SERIALIZABLE", autocommit=False)
        with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(sql.SQL("SELECT * FROM {} WHERE id=%s FOR UPDATE").format(_table(AUDIT_TABLE)), (int(audit_id),))
            audit = cursor.fetchone()
            if not audit:
                raise LookupError("Audit not found.")
            audit = dict(audit)
            if audit["status"] not in {"ready", "audited", "approved"}:
                raise ValueError("Complete column mapping before running the audit.")
            target = TARGETS[audit["target_key"]]
            aliases = _load_aliases(connection, audit["target_key"])
            database = _database_rows(connection, target)
            source = _logical_source_rows(audit, target)
            db_index = defaultdict(list)
            coarse_index = defaultdict(list)
            for row in database:
                normalized = {
                    field: _normalize_field(field, row.get(db_column), aliases)
                    for field, db_column in target["fields"].items()
                }
                row["_normalized"] = normalized
                db_index[tuple(normalized.get(field, "") for field in target["match"])].append(row)
                coarse_length = 2 if audit["target_key"] != "transection" else 3
                coarse_index[tuple(normalized.get(field, "") for field in target["match"][:coarse_length])].append(row)
            source_key_count = Counter()
            prepared = []
            for row in source:
                normalized = {field: _normalize_field(field, value, aliases) for field, value in row.items() if not field.startswith("__")}
                key = tuple(normalized.get(field, "") for field in target["match"])
                source_key_count[key] += 1
                prepared.append((row, normalized, key))
            referenced = set()
            output = []
            counts = Counter()
            for source_row, normalized, key in prepared:
                candidates = db_index.get(key, [])
                coarse_length = 2 if audit["target_key"] != "transection" else 3
                coarse_candidates = coarse_index.get(key[:coarse_length], [])
                db_row = candidates[0] if len(candidates) == 1 else None
                differences = {}
                if not candidates:
                    if coarse_candidates:
                        classification = "ambiguous"
                        referenced.update(row["id"] for row in coarse_candidates)
                        candidates = coarse_candidates
                        # Keep a canonical comparison row when the coarse key
                        # identifies exactly one candidate.  The row remains
                        # ambiguous, but Merge Alias needs the database value
                        # to persist a safe alias decision.
                        if len(coarse_candidates) == 1:
                            db_row = coarse_candidates[0]
                    else:
                        classification = "excel_only"
                elif len(candidates) > 1 or source_key_count[key] > 1:
                    classification = "probable_duplicate"
                    referenced.update(row["id"] for row in candidates)
                else:
                    referenced.add(db_row["id"])
                    for field in target["compare"]:
                        db_value = db_row.get(target["fields"][field])
                        if _number(source_row.get(field)) != _number(db_value):
                            differences[field] = {
                                "excel": _json_value(source_row.get(field)),
                                "database": _json_value(db_value),
                                "difference": _json_value((_number(source_row.get(field)) or 0) - (_number(db_value) or 0)),
                            }
                    if "quantity" in differences:
                        classification = "quantity_mismatch"
                    elif "total_amount" in differences or "amount_received" in differences:
                        classification = "amount_mismatch"
                    else:
                        alias_changed = [
                            field for field in target["alias_fields"]
                            if normalize_lookup(source_row.get(field)) != normalize_lookup(db_row.get(target["fields"][field]))
                        ]
                        classification = "voucher_normalized_match" if alias_changed else "voucher_exact_match"
                match_stage = "business_document_and_line_key" if db_row else "business_document_group"
                reason = ""
                confidence = "high" if db_row else "medium"
                if classification == "voucher_exact_match":
                    reason = "authoritative business identity and line values match"
                elif classification == "voucher_normalized_match":
                    reason = "business identity matches after approved normalization"
                elif classification == "amount_mismatch":
                    reason = "line amount differs after Decimal comparison"
                elif classification == "quantity_mismatch":
                    reason = "line quantity differs after Decimal comparison"
                elif classification == "ambiguous":
                    reason = "business document group has multiple or incomplete candidates"
                    confidence = "low"
                elif classification == "probable_duplicate":
                    reason = "logical line key occurs more times than its counterpart"
                    confidence = "medium"
                counts[classification] += 1
                differences["_diagnostics"] = {
                    "voucher_identity_key": dict(zip(target["match"], key)),
                    "matching_stage": match_stage,
                    "reason": reason,
                    "confidence": confidence,
                    "amount_interpretation": target.get("amount_interpretation", "target-defined"),
                }
                output.append((
                    int(audit_id), source_row["__row_number"], db_row["id"] if db_row else None,
                    classification, _json(dict(zip(target["match"], key))),
                    _json({key: value for key, value in source_row.items() if not key.startswith("__")}),
                    _json(normalized),
                    _json({field: db_row.get(column) for field, column in target["fields"].items()}) if db_row else "{}",
                    _json(differences), _json([row["id"] for row in candidates]),
                ))
            for row in database:
                if row["id"] not in referenced:
                    counts["database_only"] += 1
                    output.append((
                        int(audit_id), None, row["id"], "database_only",
                        _json({field: row["_normalized"].get(field) for field in target["match"]}),
                        "{}", _json(row["_normalized"]),
                        _json({field: row.get(column) for field, column in target["fields"].items()}),
                        "{}", "[]",
                    ))
            cursor.execute(sql.SQL("DELETE FROM {} WHERE audit_id=%s").format(_table(ROW_TABLE)), (int(audit_id),))
            psycopg2.extras.execute_values(
                cursor,
                sql.SQL("INSERT INTO {} (audit_id,excel_row_number,database_row_id,classification,match_key,"
                        "excel_values,normalized_values,database_values,differences,candidate_database_ids) VALUES %s").format(_table(ROW_TABLE)).as_string(connection),
                output,
                template="(%s,%s,%s,%s,%s::jsonb,%s::jsonb,%s::jsonb,%s::jsonb,%s::jsonb,%s::jsonb)",
            )
            summary = {
                "excel_rows": len(source), "database_rows": len(database),
                "exact_matches": counts["voucher_exact_match"], "alias_matches": counts["voucher_normalized_match"],
                "need_review": sum(counts[name] for name in ("ambiguous", "probable_duplicate", "quantity_mismatch", "amount_mismatch", "excel_only", "database_only")),
                "duplicates": counts["probable_duplicate"], "amount_mismatches": counts["amount_mismatch"],
                "quantity_mismatches": counts["quantity_mismatch"], "excel_only": counts["excel_only"],
                "database_only": counts["database_only"],
            }
            cursor.execute(
                sql.SQL("UPDATE {} SET status='audited',summary=%s::jsonb,rows_compared=%s,version=version+1 WHERE id=%s").format(_table(AUDIT_TABLE)),
                (_json(summary), len(output), int(audit_id)),
            )
        connection.commit()
    return get_audit(audit_id)


def list_differences(audit_id, filters):
    page = max(int(filters.get("page") or 1), 1)
    page_size = min(max(int(filters.get("page_size") or 25), 1), 100)
    clauses = ["audit_id=%(audit_id)s"]
    params = {"audit_id": int(audit_id), "limit": page_size, "offset": (page - 1) * page_size}
    if filters.get("classification"):
        clauses.append("classification=%(classification)s")
        params["classification"] = filters["classification"]
    if filters.get("decision"):
        clauses.append("decision=%(decision)s")
        params["decision"] = filters["decision"]
    if filters.get("search"):
        clauses.append("(excel_values::text ILIKE %(search)s OR database_values::text ILIKE %(search)s)")
        params["search"] = f"%{filters['search']}%"
    order = filters.get("sort") if filters.get("sort") in {"classification", "excel_row_number", "database_row_id", "decision"} else "id"
    direction = "DESC" if filters.get("direction") == "desc" else "ASC"
    where = " AND ".join(clauses)
    with _connect() as connection, connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        cursor.execute(
            sql.SQL(f"SELECT *,COUNT(*) OVER() total_count FROM {{}} WHERE {where} ORDER BY {order} {direction},id LIMIT %(limit)s OFFSET %(offset)s").format(_table(ROW_TABLE)),
            params,
        )
        rows = [dict(row) for row in cursor.fetchall()]
    total = int(rows[0].pop("total_count")) if rows else 0
    for row in rows:
        for key in ("created_at", "decided_at"):
            if row.get(key):
                row[key] = row[key].isoformat()
    return {"rows": rows, "page": page, "page_size": page_size, "total": total, "pages": max(1, (total + page_size - 1) // page_size)}


def decide(audit_id, row_ids, decision, note, username):
    allowed = {"accept_excel", "accept_database", "ignore", "merge_alias"}
    if decision not in allowed:
        raise ValueError("Invalid review action.")
    ids = [int(value) for value in row_ids]
    if not ids:
        raise ValueError("Select at least one difference.")
    with _connect() as connection, connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        if decision == "merge_alias":
            audit = get_audit(audit_id, connection)
            target = TARGETS[audit["target_key"]]
            cursor.execute(
                sql.SQL("SELECT id,excel_values,database_values FROM {} WHERE audit_id=%s AND id=ANY(%s) FOR UPDATE").format(_table(ROW_TABLE)),
                (int(audit_id), ids),
            )
            reviews = [dict(row) for row in cursor.fetchall()]
            aliases_created = 0
            for review in reviews:
                for field, alias_type in target["alias_fields"].items():
                    original = _text(review["excel_values"].get(field))
                    canonical = _text(review["database_values"].get(field))
                    if not original or not canonical or normalize_lookup(original) == normalize_lookup(canonical):
                        continue
                    cursor.execute(
                        sql.SQL("INSERT INTO {} (alias_type,original_value,normalized_value,normalized_lookup,"
                                "target_key,created_by,updated_by) VALUES (%s,%s,%s,%s,%s,%s,%s) "
                                "ON CONFLICT (alias_type,normalized_lookup,target_key) DO UPDATE SET "
                                "normalized_value=EXCLUDED.normalized_value,active=true,updated_at=now(),"
                                "updated_by=EXCLUDED.updated_by").format(_table(ALIAS_TABLE)),
                        (alias_type, original, canonical, normalize_lookup(original), audit["target_key"], username, username),
                    )
                    aliases_created += 1
            if not aliases_created:
                raise ValueError("The selected rows do not contain a customer or product alias pair.")
        cursor.execute(
            sql.SQL("UPDATE {} SET decision=%s,decision_note=%s,decided_by=%s,decided_at=now() "
                    "WHERE audit_id=%s AND id=ANY(%s) RETURNING id").format(_table(ROW_TABLE)),
            (decision, note or None, username, int(audit_id), ids),
        )
        changed = [row["id"] if isinstance(row, dict) else row[0] for row in cursor.fetchall()]
        if len(changed) != len(set(ids)):
            connection.rollback()
            raise ValueError("One or more review rows no longer exist.")
        cursor.execute(sql.SQL("UPDATE {} SET status='approved',version=version+1 WHERE id=%s").format(_table(AUDIT_TABLE)), (int(audit_id),))
        connection.commit()
    return {"updated": len(changed)}


def delete_audit(audit_id):
    with _connect() as connection, connection.cursor() as cursor:
        cursor.execute(sql.SQL("SELECT status FROM {} WHERE id=%s FOR UPDATE").format(_table(AUDIT_TABLE)), (int(audit_id),))
        row = cursor.fetchone()
        if not row:
            raise LookupError("Audit not found.")
        cursor.execute(sql.SQL("DELETE FROM {} WHERE audit_id=%s").format(_table(BACKUP_TABLE)), (int(audit_id),))
        cursor.execute(sql.SQL("DELETE FROM {} WHERE id=%s").format(_table(AUDIT_TABLE)), (int(audit_id),))
        connection.commit()
    return {"deleted": True, "audit_id": int(audit_id)}


def apply_audit(audit_id, username):
    with _connect() as connection:
        connection.set_session(isolation_level="SERIALIZABLE", autocommit=False)
        try:
            with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                cursor.execute(sql.SQL("SELECT * FROM {} WHERE id=%s FOR UPDATE").format(_table(AUDIT_TABLE)), (int(audit_id),))
                audit = cursor.fetchone()
                if not audit:
                    raise LookupError("Audit not found.")
                if audit["status"] == "applied":
                    raise ValueError("This audit was already applied.")
                target = TARGETS[audit["target_key"]]
                cursor.execute(
                    sql.SQL("SELECT * FROM {} WHERE audit_id=%s AND decision='accept_excel' ORDER BY id FOR UPDATE").format(_table(ROW_TABLE)),
                    (int(audit_id),),
                )
                approved = [dict(row) for row in cursor.fetchall()]
                applied = 0
                for review in approved:
                    excel_values = review["excel_values"]
                    if review["database_row_id"] is None:
                        importer_row = {"__row_number": review["excel_row_number"]}
                        for field, column in target["fields"].items():
                            if field in excel_values:
                                importer_row[column] = excel_values[field]
                        result = excel_importer.insert_rows_transactional(
                            connection, target["import_key"], [importer_row], commit=False
                        )
                        inserted_id = result["inserted"][0]["id"]
                        cursor.execute(
                            sql.SQL("INSERT INTO {} (audit_id,audit_row_id,target_table,database_row_id,operation,"
                                    "before_record,after_record,created_by) VALUES (%s,%s,%s,%s,'insert',NULL,%s::jsonb,%s)").format(_table(BACKUP_TABLE)),
                            (int(audit_id), review["id"], target["table"], inserted_id, _json(excel_values), username),
                        )
                    else:
                        cursor.execute(
                            sql.SQL("SELECT * FROM {} WHERE id=%s FOR UPDATE").format(_table(target["table"])),
                            (review["database_row_id"],),
                        )
                        before = cursor.fetchone()
                        if not before:
                            raise RuntimeError(f"Database row {review['database_row_id']} no longer exists.")
                        # Capture the pre-image before touching the production row.  The
                        # after-image is filled in only after the update succeeds, while
                        # the caller-owned transaction still guarantees atomic rollback.
                        cursor.execute(
                            sql.SQL("INSERT INTO {} (audit_id,audit_row_id,target_table,database_row_id,operation,"
                                    "before_record,after_record,created_by) VALUES (%s,%s,%s,%s,'update',%s::jsonb,NULL,%s) RETURNING id").format(_table(BACKUP_TABLE)),
                            (int(audit_id), review["id"], target["table"], review["database_row_id"],
                             _json(dict(before)), username),
                        )
                        backup_id = cursor.fetchone()["id"]
                        assignments, values, after = [], [], dict(before)
                        for field in target["update_fields"]:
                            if field not in excel_values:
                                continue
                            column = target["fields"][field]
                            cleaned = excel_importer.clean_value_for_column(target["import_key"], column, excel_values[field])
                            assignments.append(sql.SQL("{}=%s").format(sql.Identifier(column)))
                            values.append(cleaned)
                            after[column] = cleaned
                        if not assignments:
                            raise ValueError(f"Approved row {review['id']} has no writable fields.")
                        values.append(review["database_row_id"])
                        cursor.execute(
                            sql.SQL("UPDATE {} SET {} WHERE id=%s").format(
                                _table(target["table"]), sql.SQL(",").join(assignments)
                            ),
                            values,
                        )
                        cursor.execute(
                            sql.SQL("UPDATE {} SET after_record=%s::jsonb WHERE id=%s").format(_table(BACKUP_TABLE)),
                            (_json(after), backup_id),
                        )
                    cursor.execute(sql.SQL("UPDATE {} SET applied=true WHERE id=%s").format(_table(ROW_TABLE)), (review["id"],))
                    applied += 1
                cursor.execute(
                    sql.SQL("UPDATE {} SET status='applied',changes_applied=%s,applied_at=now(),applied_by=%s,"
                            "version=version+1 WHERE id=%s").format(_table(AUDIT_TABLE)),
                    (applied, username, int(audit_id)),
                )
            connection.commit()
            return {"audit_id": int(audit_id), "changes_applied": applied, "status": "applied"}
        except Exception:
            connection.rollback()
            raise


def list_history(filters):
    page = max(int(filters.get("page") or 1), 1)
    page_size = min(max(int(filters.get("page_size") or 20), 1), 100)
    params = {"limit": page_size, "offset": (page - 1) * page_size}
    with _connect() as connection, connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        cursor.execute(
            sql.SQL("SELECT id,created_at,created_by,target_key,target_table,filename,source_rows,rows_compared,"
                    "changes_applied,warnings,status,COUNT(*) OVER() total_count FROM {} "
                    "ORDER BY created_at DESC,id DESC LIMIT %(limit)s OFFSET %(offset)s").format(_table(AUDIT_TABLE)),
            params,
        )
        rows = [dict(row) for row in cursor.fetchall()]
    total = int(rows[0].pop("total_count")) if rows else 0
    for row in rows:
        row["created_at"] = row["created_at"].isoformat()
    return {"audits": rows, "page": page, "page_size": page_size, "total": total, "pages": max(1, (total + page_size - 1) // page_size)}


def list_aliases(alias_type="", search=""):
    clauses, params = ["1=1"], {}
    if alias_type:
        clauses.append("alias_type=%(alias_type)s")
        params["alias_type"] = alias_type
    if search:
        clauses.append("(original_value ILIKE %(search)s OR normalized_value ILIKE %(search)s)")
        params["search"] = f"%{search}%"
    with _connect() as connection, connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        cursor.execute(
            sql.SQL("SELECT * FROM {} WHERE {} ORDER BY alias_type,original_value").format(
                _table(ALIAS_TABLE), sql.SQL(" AND ").join(sql.SQL(clause) for clause in clauses)
            ), params,
        )
        return [dict(row) for row in cursor.fetchall()]


def save_alias(alias_id, payload, username):
    alias_type = _text(payload.get("alias_type"))
    original = _text(payload.get("original_value"))
    normalized = _text(payload.get("normalized_value"))
    target_key = _text(payload.get("target_key")) or None
    if alias_type not in {"customer", "product"} or not original or not normalized:
        raise ValueError("Alias type, original value and normalized value are required.")
    if target_key and target_key not in TARGETS:
        raise ValueError("Unsupported alias target.")
    with _connect() as connection, connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        if alias_id:
            cursor.execute(
                sql.SQL("UPDATE {} SET alias_type=%s,original_value=%s,normalized_value=%s,normalized_lookup=%s,"
                        "target_key=%s,active=%s,updated_at=now(),updated_by=%s WHERE id=%s RETURNING *").format(_table(ALIAS_TABLE)),
                (alias_type, original, normalized, normalize_lookup(original), target_key,
                 bool(payload.get("active", True)), username, int(alias_id)),
            )
        else:
            cursor.execute(
                sql.SQL("INSERT INTO {} (alias_type,original_value,normalized_value,normalized_lookup,target_key,"
                        "created_by,updated_by) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *").format(_table(ALIAS_TABLE)),
                (alias_type, original, normalized, normalize_lookup(original), target_key, username, username),
            )
        row = cursor.fetchone()
        connection.commit()
        return dict(row)


def export_csv(audit_id, mode="all"):
    filters = {"page": 1, "page_size": 100}
    rows = []
    page = 1
    while True:
        filters["page"] = page
        result = list_differences(audit_id, filters)
        rows.extend(result["rows"])
        if page >= result["pages"]:
            break
        page += 1
    if mode == "mismatch":
        rows = [row for row in rows if row["classification"] not in {"exact_match", "alias_match", "voucher_exact_match", "voucher_normalized_match"}]
    output = io.StringIO()
    fields = ["id", "excel_row_number", "database_row_id", "classification", "decision",
              "excel_values", "database_values", "differences", "voucher_identity_key",
              "matching_stage", "reason", "confidence", "amount_interpretation",
              "decision_note", "decided_by", "applied"]
    writer = csv.DictWriter(output, fieldnames=fields)
    writer.writeheader()
    for row in rows:
        diagnostics = (row.get("differences") or {}).get("_diagnostics") or {}
        values = {key: _json(row[key]) if isinstance(row.get(key), (dict, list)) else row.get(key, "") for key in fields}
        values.update({
            "voucher_identity_key": _json(diagnostics.get("voucher_identity_key", {})),
            "matching_stage": diagnostics.get("matching_stage", ""),
            "reason": diagnostics.get("reason", ""),
            "confidence": diagnostics.get("confidence", ""),
            "amount_interpretation": diagnostics.get("amount_interpretation", ""),
        })
        writer.writerow(values)
    return output.getvalue()


def export_change_log(audit_id):
    with _connect() as connection, connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        cursor.execute(sql.SQL("SELECT * FROM {} WHERE audit_id=%s ORDER BY id").format(_table(BACKUP_TABLE)), (int(audit_id),))
        rows = [dict(row) for row in cursor.fetchall()]
    output = io.StringIO()
    fields = ["id", "audit_row_id", "target_table", "database_row_id", "operation", "before_record", "after_record", "created_at", "created_by"]
    writer = csv.DictWriter(output, fieldnames=fields)
    writer.writeheader()
    for row in rows:
        writer.writerow({key: _json(row[key]) if isinstance(row.get(key), (dict, list)) else _text(row.get(key)) for key in fields})
    return output.getvalue()


def export_summary_pdf(audit_id, output_path):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    audit = get_audit(audit_id)
    document = canvas.Canvas(str(output_path), pagesize=A4)
    width, height = A4
    y = height - 54
    document.setFont("Helvetica-Bold", 16)
    document.drawString(45, y, f"BigShot Data Audit #{audit_id}")
    y -= 28
    document.setFont("Helvetica", 10)
    for label, value in (
        ("Target", audit["target_table"]), ("File", audit["filename"]), ("Status", audit["status"]),
        ("Created by", audit["created_by"]), ("Rows compared", audit["rows_compared"]),
        ("Changes applied", audit["changes_applied"]),
    ):
        document.drawString(45, y, f"{label}: {_text(value)}")
        y -= 17
    y -= 10
    document.setFont("Helvetica-Bold", 12)
    document.drawString(45, y, "Summary")
    document.setFont("Helvetica", 10)
    for key, value in audit["summary"].items():
        y -= 17
        document.drawString(55, y, f"{key.replace('_', ' ').title()}: {_text(value)}")
    document.setFont("Helvetica", 8)
    document.drawString(45, 30, "Generated by BigShot AI Business OS Data Audit Center")
    document.save()
