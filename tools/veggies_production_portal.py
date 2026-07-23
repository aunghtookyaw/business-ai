"""Browser-first Veggies Production Basic routes and database operations."""

from __future__ import annotations

import json
import io
import uuid
from collections import OrderedDict
from datetime import date
from decimal import Decimal
from html import escape
from typing import Any
from urllib.parse import urlencode

import psycopg2.extras
from flask import redirect, request, send_file, url_for
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from tools import formula_engine
from tools.veggies_production import (
    CROP_CATEGORIES,
    CropDefinition,
    load_crop_definitions,
    format_quantity,
    parse_production_date,
    parse_quantity,
)


PAGE_SIZE = 25


def _ref(table: str) -> str:
    return f'"public"."{table}"'


def _connect():
    return formula_engine._connect()


def _text(value: Any) -> str:
    return str(value or "").strip()


def _date_text(value: Any) -> str:
    if isinstance(value, (date,)):
        return value.isoformat()
    return _text(value)


def _quantity_text(value: Any) -> str:
    return format_quantity(value)


def _quantity_input_text(value: Any) -> str:
    """Apply the shared display precision to browser quantity fields."""
    return format_quantity(value)


def _quantity_storage_text(value: Any) -> str:
    """Serialize untouched input precision so an edit-only save cannot round the database value."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return ""
    return format(Decimal(str(value)), "f")


def _crop_field(crop: CropDefinition) -> str:
    return f"crop_{crop.crop_code}"


def portal_crops() -> list[CropDefinition]:
    """Return one active browser field per crop while importer aliases remain intact."""
    unique = {}
    for crop in load_crop_definitions():
        unique.setdefault(crop.crop_code, crop)
    return list(unique.values())


def portal_farm_areas(connection=None) -> list[dict[str, Any]]:
    """Load active farm areas dynamically for production entry and filtering."""
    owns_connection = connection is None
    connection = connection or _connect()
    try:
        with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(f"""
                SELECT id, area_code, area_name
                FROM {_ref('veggies_farm_area_master')}
                WHERE active = TRUE
                ORDER BY display_order, area_name
            """)
            return [dict(row) for row in cursor.fetchall()]
    finally:
        if owns_connection:
            connection.close()


def grouped_crops(crops: list[CropDefinition]) -> OrderedDict[str, list[CropDefinition]]:
    """Group crops in the stable staff-facing category order."""
    groups = OrderedDict((category, []) for category in CROP_CATEGORIES)
    for crop in crops:
        groups.setdefault(crop.category or "Other", []).append(crop)
    return OrderedDict((category, rows) for category, rows in groups.items() if rows)


def validate_submission(values: dict[str, Any], crops: list[CropDefinition],
                        farm_areas: list[dict[str, Any]] | None = None) -> tuple[dict[str, Any], dict[str, str]]:
    """Validate a browser submission while preserving blank versus zero."""
    errors: dict[str, str] = {}
    valid_area_ids = {int(area["id"]) for area in (farm_areas or [])}
    try:
        farm_area_id = int(values.get("farm_area_id") or 0)
    except (TypeError, ValueError):
        farm_area_id = 0
    if not farm_area_id or farm_area_id not in valid_area_ids:
        errors["farm_area_id"] = "Select a valid Farm Area."
    try:
        production_date = parse_production_date(values.get("production_date"))
        if production_date is None:
            raise ValueError("Production Date is required.")
    except ValueError as exc:
        production_date = None
        errors["production_date"] = str(exc)

    entry_text = _text(values.get("entry_date"))
    try:
        entry_date = parse_production_date(entry_text, "Date of Entry") if entry_text else date.today()
    except ValueError as exc:
        entry_date = None
        errors["entry_date"] = str(exc)

    items = []
    for crop in crops:
        field = _crop_field(crop)
        try:
            quantity = parse_quantity(values.get(field), crop.crop_name)
        except ValueError as exc:
            errors[field] = str(exc)
            continue
        if quantity is not None:
            items.append({
                "crop_id": crop.crop_id,
                "crop_code": crop.crop_code,
                "crop_name": crop.crop_name,
                "quantity": quantity,
                "unit": crop.default_unit,
            })
    if not items and not any(key.startswith("crop_") for key in errors):
        errors["crops"] = "Enter at least one vegetable quantity. Zero is allowed."

    return {
        "production_date": production_date,
        "farm_area_id": farm_area_id or None,
        "assignee": _text(values.get("assignee")) or None,
        "note": _text(values.get("note")) or None,
        "ai_note": _text(values.get("ai_note")) or None,
        "entry_date": entry_date,
        "submission_token": _text(values.get("submission_token")) or str(uuid.uuid4()),
        "items": items,
    }, errors


def save_submission(values: dict[str, Any], connection=None) -> dict[str, Any]:
    """Insert one batch and its normalized items in a single transaction."""
    owns_connection = connection is None
    connection = connection or _connect()
    try:
        with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(
                f"""
                INSERT INTO {_ref('veggies_production_batches')}
                  (production_date, farm_area_id, assignee, note, ai_note, entry_date,
                   submission_token, created_by, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                ON CONFLICT (submission_token) WHERE submission_token IS NOT NULL DO NOTHING
                RETURNING id
                """,
                (values["production_date"], values["farm_area_id"], values["assignee"], values["note"],
                 values["ai_note"], values["entry_date"], values["submission_token"],
                 "Veggies Production Basic"),
            )
            row = cursor.fetchone()
            if not row:
                raise ValueError("This submission was already saved. Refresh before entering another record.")
            batch_id = row["id"]
            for item in values["items"]:
                cursor.execute(
                    f"""
                    INSERT INTO {_ref('veggies_production_items')}
                      (production_batch_id, crop_id, quantity, unit, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, NOW(), NOW())
                    """,
                    (batch_id, item["crop_id"], item["quantity"], item["unit"]),
                )
        connection.commit()
        return {"id": batch_id}
    except Exception:
        connection.rollback()
        raise
    finally:
        if owns_connection:
            connection.close()


def update_submission(batch_id: int, values: dict[str, Any], connection=None) -> dict[str, Any]:
    """Explicitly replace a batch's editable values and items in one transaction."""
    owns_connection = connection is None
    connection = connection or _connect()
    try:
        with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(
                f"SELECT id FROM {_ref('veggies_production_batches')} WHERE id = %s FOR UPDATE",
                (batch_id,),
            )
            if not cursor.fetchone():
                raise LookupError("Production record not found.")
            cursor.execute(
                f"""
                UPDATE {_ref('veggies_production_batches')}
                SET production_date=%s, farm_area_id=%s, assignee=%s, note=%s, ai_note=%s,
                    entry_date=%s, updated_at=NOW()
                WHERE id=%s
                """,
                (values["production_date"], values["farm_area_id"], values["assignee"], values["note"],
                 values["ai_note"], values["entry_date"], batch_id),
            )
            cursor.execute(
                f"DELETE FROM {_ref('veggies_production_items')} WHERE production_batch_id = %s",
                (batch_id,),
            )
            for item in values["items"]:
                cursor.execute(
                    f"""
                    INSERT INTO {_ref('veggies_production_items')}
                      (production_batch_id, crop_id, quantity, unit, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, NOW(), NOW())
                    """,
                    (batch_id, item["crop_id"], item["quantity"], item["unit"]),
                )
        connection.commit()
        return {"id": batch_id}
    except Exception:
        connection.rollback()
        raise
    finally:
        if owns_connection:
            connection.close()


def delete_submission(batch_id: int, confirmation: Any, reason: Any, connection=None) -> dict[str, Any]:
    """Delete one manually entered production batch and its items atomically."""
    if _text(confirmation) != str(int(batch_id)):
        raise ValueError("Production record ID confirmation does not match")
    reason = _text(reason)
    if not reason:
        raise ValueError("Deletion reason is required")
    owns_connection = connection is None
    connection = connection or _connect()
    try:
        with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(f"SELECT * FROM {_ref('veggies_production_batches')} WHERE id=%s FOR UPDATE", (int(batch_id),))
            batch = cursor.fetchone()
            if not batch:
                raise LookupError("Production record not found.")
            if batch.get("import_id") is not None:
                raise ValueError("Imported production records cannot be deleted here")
            cursor.execute(f"DELETE FROM {_ref('veggies_production_items')} WHERE production_batch_id=%s", (int(batch_id),))
            cursor.execute(f"DELETE FROM {_ref('veggies_production_batches')} WHERE id=%s RETURNING id", (int(batch_id),))
            if not cursor.fetchone():
                raise RuntimeError("Production record changed concurrently; refresh and retry")
        connection.commit()
        return {"deleted": True, "batch_id": int(batch_id), "reason": reason}
    except Exception:
        connection.rollback()
        raise
    finally:
        if owns_connection:
            connection.close()


def search_records(filters: dict[str, str]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    conditions = ["TRUE"]
    params: list[Any] = []
    mappings = [
        ("date_from", "batch.production_date >= %s"),
        ("date_to", "batch.production_date <= %s"),
        ("production_date", "batch.production_date = %s"),
    ]
    for key, condition in mappings:
        if filters.get(key):
            conditions.append(condition)
            params.append(filters[key])
    if filters.get("assignee"):
        conditions.append("COALESCE(batch.assignee, '') ILIKE %s")
        params.append(f"%{filters['assignee']}%")
    if filters.get("farm_area_id"):
        conditions.append("batch.farm_area_id = %s")
        params.append(filters["farm_area_id"])
    if filters.get("crop"):
        conditions.append("crop.crop_code = %s")
        params.append(filters["crop"])
    if filters.get("min_quantity"):
        conditions.append("item.quantity >= %s")
        params.append(filters["min_quantity"])
    if filters.get("max_quantity"):
        conditions.append("item.quantity <= %s")
        params.append(filters["max_quantity"])
    if filters.get("note"):
        conditions.append("COALESCE(batch.note, '') ILIKE %s")
        params.append(f"%{filters['note']}%")
    where_sql = " AND ".join(conditions)
    sort_sql = {
        "newest": "production_date DESC, id DESC",
        "oldest": "production_date ASC, id ASC",
        "highest": "total_quantity DESC, production_date DESC, id DESC",
        "lowest": "total_quantity ASC, production_date DESC, id DESC",
    }.get(filters.get("sort"), "production_date DESC, id DESC")
    try:
        page = max(int(filters.get("page") or 1), 1)
    except ValueError:
        page = 1
    base_sql = f"""
      SELECT batch.id, batch.production_date, batch.farm_area_id, area.area_name AS farm_area,
             batch.assignee, batch.note, batch.entry_date,
             SUM(item.quantity) AS total_quantity, COUNT(DISTINCT item.crop_id) AS crop_count,
             batch.created_at, batch.updated_at
      FROM {_ref('veggies_production_batches')} batch
      JOIN {_ref('veggies_production_items')} item ON item.production_batch_id = batch.id
      JOIN {_ref('veggies_crop_master')} crop ON crop.id = item.crop_id
      JOIN {_ref('veggies_farm_area_master')} area ON area.id = batch.farm_area_id
      WHERE {where_sql}
      GROUP BY batch.id, batch.production_date, batch.farm_area_id, area.area_name,
               batch.assignee, batch.note, batch.entry_date,
               batch.created_at, batch.updated_at
    """
    connection = _connect()
    try:
        with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(
                f"SELECT *, COUNT(*) OVER() AS filtered_count FROM ({base_sql}) filtered ORDER BY {sort_sql} LIMIT %s OFFSET %s",
                [*params, PAGE_SIZE, (page - 1) * PAGE_SIZE],
            )
            rows = [dict(row) for row in cursor.fetchall()]
            total_records = int(rows[0]["filtered_count"]) if rows else 0
        summaries = {"total_records": total_records, "page": page,
                     "total_pages": max((total_records + PAGE_SIZE - 1) // PAGE_SIZE, 1)}
        return rows, summaries
    finally:
        connection.close()


def today_summary(connection=None) -> dict[str, Any]:
    """Return quantities saved today using the database's configured local date."""
    owns_connection = connection is None
    connection = connection or _connect()
    try:
        with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(f"""
                SELECT COALESCE(SUM(item.quantity), 0) AS total_quantity,
                       COUNT(DISTINCT batch.id) AS submission_count,
                       COUNT(DISTINCT item.crop_id) AS crop_count,
                       MAX(batch.created_at) AS latest_entry_time,
                       COALESCE(BOOL_OR(item.unit IS NULL), TRUE)
                         OR COUNT(DISTINCT item.unit) > 1 AS unit_pending
                FROM {_ref('veggies_production_batches')} batch
                LEFT JOIN {_ref('veggies_production_items')} item
                  ON item.production_batch_id = batch.id
                WHERE batch.created_at::date = CURRENT_DATE
            """)
            return dict(cursor.fetchone())
    finally:
        if owns_connection:
            connection.close()


def list_crop_master() -> list[dict[str, Any]]:
    connection = _connect()
    try:
        with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(f"""
                SELECT id, crop_code, crop_name, category, active, default_unit, display_order,
                       created_at, updated_at
                FROM {_ref('veggies_crop_master')}
                ORDER BY display_order, crop_name
            """)
            return [dict(row) for row in cursor.fetchall()]
    finally:
        connection.close()


def update_crop_master(crop_id: int, values: dict[str, Any], connection=None) -> None:
    """Update editable master fields; crops are deactivated, never deleted."""
    crop_name = _text(values.get("crop_name"))
    category = _text(values.get("category")) or "Other"
    if not crop_name:
        raise ValueError("Crop Name is required.")
    if category not in CROP_CATEGORIES:
        raise ValueError("Select a valid category.")
    try:
        display_order = int(values.get("display_order") or 0)
    except (TypeError, ValueError) as exc:
        raise ValueError("Display Order must be a whole number.") from exc
    owns_connection = connection is None
    connection = connection or _connect()
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"""
                UPDATE {_ref('veggies_crop_master')}
                SET crop_name=%s, crop_name_normalized=%s, category=%s, active=%s,
                    default_unit=%s, display_order=%s, updated_at=NOW()
                WHERE id=%s
            """, (crop_name, crop_name.casefold(), category, values.get("active") == "yes",
                  _text(values.get("default_unit")) or None, display_order, crop_id))
            if cursor.rowcount != 1:
                raise LookupError("Crop not found.")
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        if owns_connection:
            connection.close()


def get_record(batch_id: int) -> dict[str, Any] | None:
    connection = _connect()
    try:
        with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(
                f"""SELECT batch.*, area.area_name AS farm_area
                    FROM {_ref('veggies_production_batches')} batch
                    JOIN {_ref('veggies_farm_area_master')} area ON area.id=batch.farm_area_id
                    WHERE batch.id=%s""",
                (batch_id,),
            )
            batch = cursor.fetchone()
            if not batch:
                return None
            cursor.execute(
                f"""
                SELECT item.crop_id, crop.crop_code, crop.crop_name, crop.category,
                       item.quantity, item.unit,
                       item.created_at, item.updated_at
                FROM {_ref('veggies_production_items')} item
                JOIN {_ref('veggies_crop_master')} crop ON crop.id=item.crop_id
                WHERE item.production_batch_id=%s
                ORDER BY crop.display_order, crop.crop_name
                """,
                (batch_id,),
            )
            result = dict(batch)
            result["items"] = [dict(row) for row in cursor.fetchall()]
            return result
    finally:
        connection.close()


BASE_STYLE = """
body{margin:0;font-family:Arial,sans-serif;background:#f4f6f5;color:#17211c}header{padding:18px 22px;background:#fff;border-bottom:1px solid #ccd5d0}main{padding:16px;max-width:1600px;margin:auto}section{background:#fff;border:1px solid #d5ddd9;border-radius:8px;margin-bottom:16px;padding:16px}.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px}.crop-category{border:1px solid #dce4e0;border-radius:7px;margin:12px 0;padding:12px}.crop-category h4{margin:0 0 10px;color:#174f3b}.crop-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:10px;padding:4px}.crop-tools{display:grid;grid-template-columns:minmax(220px,1fr) 220px;gap:12px;align-items:end}.crop-field.hidden,.crop-category.hidden{display:none}.cards{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:16px}.card{background:#174f3b;color:white;border-radius:8px;padding:16px}.card strong{display:block;font-size:24px;margin-top:6px}.card small{display:block;margin-top:8px;color:#e4f0ea}label{display:block;font-weight:bold;font-size:12px;color:#47554e}input,textarea,select,button{box-sizing:border-box;width:100%;min-height:42px;padding:9px;margin-top:5px;border:1px solid #aebbb5;border-radius:5px;font:inherit}button,.button{background:#176b5d;color:white;border:0;font-weight:bold;cursor:pointer;text-decoration:none;display:inline-block;padding:11px 16px;border-radius:5px;width:auto}.secondary{background:#fff;color:#176b5d;border:1px solid #176b5d}.error{color:#b00020;font-size:12px;margin-top:4px}.status{padding:14px;border-radius:6px;background:#e1f3ed;color:#155d4f;font-weight:bold;line-height:1.5}.status.bad{background:#fde8e8;color:#9b1c1c}.preview{background:#f7faf8;border-left:4px solid #176b5d}.preview ul{columns:2;padding-left:20px}.table-wrap{overflow:auto}table{width:100%;min-width:1050px;border-collapse:collapse}th,td{padding:9px;border-bottom:1px solid #e1e6e3;text-align:left;white-space:nowrap}th{background:#edf2ef}.actions{display:flex;gap:8px;flex-wrap:wrap;margin-top:14px}.required{color:#a11}.original,.warning{background:#fff8e6;padding:12px;border-left:4px solid #d6a84b;margin-bottom:14px}.pagination{display:flex;align-items:center;gap:10px;margin-top:14px}@media(max-width:800px){.cards{grid-template-columns:1fr 1fr}.crop-grid{grid-template-columns:repeat(2,1fr)}.crop-tools{grid-template-columns:1fr}}@media(max-width:520px){.cards,.crop-grid{grid-template-columns:1fr}.preview ul{columns:1}}
"""


def _error(errors: dict[str, str], field: str) -> str:
    return f'<div class="error">{escape(errors[field])}</div>' if field in errors else ""


def _entry_form(crops, farm_areas, values, errors, action, edit=False):
    item_values = values.get("item_values", {})
    categories = []
    for category, category_crops in grouped_crops(crops).items():
        fields = []
        for crop in category_crops:
            field = _crop_field(crop)
            value = values.get(field, item_values.get(crop.crop_code, ""))
            input_id = f"input_{field}"
            fields.append(f'''<div class="crop-field" data-crop-name="{escape(crop.crop_name.casefold())}"><label for="{input_id}">{escape(crop.crop_name)}</label><input class="crop-input" id="{input_id}" type="number" step="any" min="0" name="{field}" value="{escape(_quantity_input_text(value))}" inputmode="decimal" data-stored-value="{escape(_quantity_storage_text(value))}" data-crop-label="{escape(crop.crop_name)}">{_error(errors, field)}</div>''')
        categories.append(f'''<div class="crop-category" data-category="{escape(category)}"><h4>{escape(category)}</h4><div class="crop-grid">{''.join(fields)}</div></div>''')
    confirm = '<label><input style="width:auto;min-height:auto" type="checkbox" name="confirm_changes" value="yes" required> I reviewed the original values and confirm these changes.</label>' if edit else ""
    selected_area = _text(values.get("farm_area_id"))
    area_options = ''.join(
        f'<option value="{area["id"]}" {"selected" if selected_area == str(area["id"]) else ""}>{escape(area["area_name"])}</option>'
        for area in farm_areas
    )
    return f'''<form method="post" action="{escape(action)}" id="productionForm">
      <input type="hidden" name="submission_token" value="{escape(_text(values.get('submission_token')) or str(uuid.uuid4()))}">
      <div class="grid">
        <label>Production Date <span class="required">*</span><input type="date" name="production_date" value="{escape(_date_text(values.get('production_date')))}" required>{_error(errors,'production_date')}</label>
        <label>Farm Area <span class="required">*</span><select name="farm_area_id" required><option value="">Select Farm Area</option>{area_options}</select>{_error(errors,'farm_area_id')}</label>
        <label>Assignee<input name="assignee" value="{escape(_text(values.get('assignee')))}"></label>
        <label>Date of Entry<input type="date" name="entry_date" value="{escape(_date_text(values.get('entry_date')) or date.today().isoformat())}">{_error(errors,'entry_date')}</label>
      </div>
      <h3>Vegetable quantities <span class="required">*</span></h3>{_error(errors,'crops')}
      <div class="crop-tools"><label for="cropSearch">Search crop<input id="cropSearch" type="search" placeholder="Example: tomato" autocomplete="off"></label><label for="cropMode">Show<select id="cropMode"><option value="all">All Crops</option><option value="entered">Entered Crops Only</option></select></label></div>
      <div id="cropSections">{''.join(categories)}</div>
      <div class="grid">
        <label>Note<textarea name="note" rows="3">{escape(_text(values.get('note')))}</textarea></label>
        <label>AI Note<textarea name="ai_note" rows="3">{escape(_text(values.get('ai_note')))}</textarea></label>
      </div>
      <section class="preview" aria-live="polite"><h3>Entry preview</h3><div class="grid"><p><b>Production Date</b><br><span id="previewDate">—</span></p><p><b>Assignee</b><br><span id="previewAssignee">—</span></p><p><b>Number of Entered Crops</b><br><span id="previewCount">0</span></p><p><b>Total Entered Quantity</b><br><span id="previewTotal">0.00</span></p></div><ul id="previewItems"><li>No crop quantities entered.</li></ul></section>{confirm}
      <div class="actions"><button id="saveButton" type="submit">{'Save Changes' if edit else 'Save Production'}</button></div>
    </form><script>
    (()=>{{
      const form=document.getElementById('productionForm'),inputs=[...form.querySelectorAll('.crop-input')],search=document.getElementById('cropSearch'),mode=document.getElementById('cropMode'),dateInput=form.querySelector('[name="production_date"]'),assigneeInput=form.querySelector('[name="assignee"]');
      const entered=input=>input.value.trim()!=='';
      function safe(text){{return text.replace(/&/g,'&amp;').replace(/</g,'&lt;');}}
      const formatQuantity=value=>(Number(value)||0).toFixed(2);
      function refresh(){{const query=search.value.trim().toLowerCase();let count=0,total=0,items=[];inputs.forEach(input=>{{const field=input.closest('.crop-field'),isEntered=entered(input),matches=!query||field.dataset.cropName.includes(query),visible=matches&&(mode.value==='all'||isEntered);field.classList.toggle('hidden',!visible);if(isEntered){{count++;const quantity=Number(input.value)||0;total+=quantity;items.push(`${{input.dataset.cropLabel}} — ${{formatQuantity(quantity)}}`);}}}});document.querySelectorAll('.crop-category').forEach(section=>section.classList.toggle('hidden',![...section.querySelectorAll('.crop-field')].some(field=>!field.classList.contains('hidden'))));document.getElementById('previewDate').textContent=dateInput.value||'—';document.getElementById('previewAssignee').textContent=assigneeInput.value.trim()||'—';document.getElementById('previewCount').textContent=count;document.getElementById('previewTotal').textContent=formatQuantity(total);document.getElementById('previewItems').innerHTML=items.length?items.map(item=>`<li>${{safe(item)}}</li>`).join(''):'<li>No crop quantities entered.</li>';}}
      inputs.forEach(input=>{{input.addEventListener('input',()=>{{input.dataset.changed='1';refresh();}});input.addEventListener('focus',()=>{{if(input.dataset.changed!=='1'&&entered(input))input.value=input.dataset.storedValue;}});input.addEventListener('blur',()=>{{if(entered(input))input.value=formatQuantity(input.value);refresh();}});}});search.addEventListener('input',refresh);mode.addEventListener('change',refresh);dateInput.addEventListener('input',refresh);assigneeInput.addEventListener('input',refresh);refresh();form.addEventListener('submit',()=>{{inputs.forEach(input=>{{if(input.dataset.changed!=='1')input.value=input.dataset.storedValue;}});const b=document.getElementById('saveButton');b.disabled=true;b.textContent='Saving…';}});
    }})();
    </script>'''


def _record_pdf(record: dict[str, Any]) -> io.BytesIO:
    output = io.BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    pdf.setTitle(f"Veggies Production {record['id']}")
    y = A4[1] - 48
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(42, y, "Veggies Production Record")
    pdf.setFont("Helvetica", 10)
    for text in (
        f"Record: {record['id']}", f"Production date: {_date_text(record['production_date'])}",
        f"Farm area: {_text(record['farm_area'])}", f"Assignee: {_text(record['assignee']) or '-'}",
    ):
        y -= 20
        pdf.drawString(42, y, text)
    y -= 28
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(42, y, "Crop")
    pdf.drawString(300, y, "Quantity")
    pdf.drawString(390, y, "Unit")
    pdf.setFont("Helvetica", 10)
    for item in record["items"]:
        y -= 18
        pdf.drawString(42, y, _text(item["crop_name"])[:38])
        pdf.drawRightString(365, y, format_quantity(item["quantity"]))
        pdf.drawString(390, y, _text(item["unit"]) or "-")
    pdf.save()
    output.seek(0)
    return output


def _record_excel(record: dict[str, Any]) -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Production Record"
    sheet.append(["Record ID", record["id"]])
    sheet.append(["Production Date", record["production_date"]])
    sheet.append(["Farm Area", record["farm_area"]])
    sheet.append(["Assignee", record["assignee"] or ""])
    sheet.append([])
    sheet.append(["Crop", "Quantity", "Unit"])
    for item in record["items"]:
        sheet.append([item["crop_name"], item["quantity"], item["unit"] or ""])
        sheet.cell(sheet.max_row, 2).number_format = "0.00"
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _filters() -> dict[str, str]:
    return {key: _text(request.args.get(key)) for key in (
        "date_from", "date_to", "production_date", "assignee", "crop",
        "farm_area_id", "min_quantity", "max_quantity", "note", "sort", "page",
    )}


def _render_main(crops, farm_areas, values=None, errors=None, message="", bad=False):
    values = values or {}
    errors = errors or {}
    filters = _filters()
    try:
        rows, summary = search_records(filters)
        load_error = ""
    except Exception as exc:
        rows, summary = [], {"total_records": 0, "page": 1, "total_pages": 1}
        load_error = f"Could not load production records: {exc}"
    try:
        today = today_summary()
    except Exception as exc:
        today = {"total_quantity": 0, "submission_count": 0, "crop_count": 0,
                 "latest_entry_time": None, "unit_pending": True}
        load_error = load_error or f"Could not load today’s summary: {exc}"
    status = message or load_error
    status_html = f'<p class="status {"bad" if bad or load_error else ""}">{escape(status)}</p>' if status else ""
    crop_options = ''.join(f'<option value="{escape(c.crop_code)}" {"selected" if filters["crop"]==c.crop_code else ""}>{escape(c.crop_name)}</option>' for c in crops)
    area_filter_options = ''.join(f'<option value="{area["id"]}" {"selected" if filters["farm_area_id"]==str(area["id"]) else ""}>{escape(area["area_name"])}</option>' for area in farm_areas)
    result_rows = ''.join(
        f'<tr><td>{escape(_date_text(row["production_date"]))}</td><td>{escape(_text(row["farm_area"]))}</td><td>{escape(_text(row["assignee"]))}</td><td>{escape(_quantity_text(row["total_quantity"]))}</td><td>{row["crop_count"]}</td><td>{escape(_text(row["note"]))}</td><td>{escape(_date_text(row["entry_date"]))}</td><td><a href="/veggies-production/{row["id"]}">View</a></td><td><a href="/veggies-production/{row["id"]}/edit">Edit</a></td></tr>'
        for row in rows
    ) or '<tr><td colspan="9">No production records found.</td></tr>'
    query_base = {key: value for key, value in filters.items() if value and key != "page"}
    previous_link = ""
    next_link = ""
    if summary["page"] > 1:
        previous_link = f'<a class="button secondary" href="/veggies-production?{urlencode({**query_base, "page": summary["page"] - 1})}">Previous</a>'
    if summary["page"] < summary["total_pages"]:
        next_link = f'<a class="button secondary" href="/veggies-production?{urlencode({**query_base, "page": summary["page"] + 1})}">Next</a>'
    latest_time = today["latest_entry_time"].strftime("%H:%M:%S") if hasattr(today["latest_entry_time"], "strftime") else "—"
    unit_note = '<small>Unit configuration pending</small>' if today.get("unit_pending") else ""
    return f'''<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Veggies Production Basic</title><style>{BASE_STYLE}</style></head><body>
    <header><h1>Veggies Production Basic</h1></header><main>{status_html}
    <div class="cards"><div class="card">Today’s Total Production<strong>{escape(_quantity_text(today['total_quantity']))}</strong>{unit_note}</div><div class="card">Today’s Number of Submissions<strong>{today['submission_count']}</strong></div><div class="card">Today’s Number of Crops Produced<strong>{today['crop_count']}</strong></div><div class="card">Latest Entry Time<strong>{escape(latest_time)}</strong></div></div>
    <section><div class="actions"><h2 style="margin-right:auto">New production submission</h2><a class="button secondary" href="/veggies-production/crops">Veggies Crop Master</a></div>{_entry_form(crops, farm_areas, values, errors, '/veggies-production')}</section>
    <section><h2>Search and review</h2><form method="get" action="/veggies-production"><div class="grid">
      <label>Date From<input type="date" name="date_from" value="{escape(filters['date_from'])}"></label><label>Date To<input type="date" name="date_to" value="{escape(filters['date_to'])}"></label><label>Production Date<input type="date" name="production_date" value="{escape(filters['production_date'])}"></label><label>Farm Area<select name="farm_area_id"><option value="">All areas</option>{area_filter_options}</select></label><label>Assignee<input name="assignee" value="{escape(filters['assignee'])}"></label><label>Crop<select name="crop"><option value="">All crops</option>{crop_options}</select></label><label>Minimum Quantity<input type="number" step="any" min="0" name="min_quantity" value="{escape(filters['min_quantity'])}"></label><label>Maximum Quantity<input type="number" step="any" min="0" name="max_quantity" value="{escape(filters['max_quantity'])}"></label><label>Note search<input name="note" value="{escape(filters['note'])}"></label><label>Sort<select name="sort"><option value="newest" {"selected" if filters['sort'] in ('','newest') else ''}>Newest first</option><option value="oldest" {"selected" if filters['sort']=='oldest' else ''}>Oldest first</option><option value="highest" {"selected" if filters['sort']=='highest' else ''}>Highest total quantity</option><option value="lowest" {"selected" if filters['sort']=='lowest' else ''}>Lowest total quantity</option></select></label></div><div class="actions"><button type="submit">Search</button><a class="button secondary" href="/veggies-production">Clear</a></div></form>
      <p>{summary['total_records']} record(s)</p><div class="table-wrap"><table><thead><tr><th>Production Date</th><th>Farm Area</th><th>Assignee</th><th>Total Quantity</th><th>Number of Crops</th><th>Note</th><th>Date of Entry</th><th>View</th><th>Edit</th></tr></thead><tbody>{result_rows}</tbody></table></div><div class="pagination">{previous_link}<span>Page {summary['page']} of {summary['total_pages']}</span>{next_link}</div></section>
    </main></body></html>''', 200, {"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"}


def register_routes(app) -> None:
    @app.route("/veggies-production", methods=["GET", "POST"])
    def veggies_production_basic():
        try:
            crops = portal_crops()
            farm_areas = portal_farm_areas()
        except Exception as exc:
            return _render_main([], [], message=f"Could not load production master data: {exc}", bad=True)
        if request.method == "GET":
            message = ""
            if request.args.get("deleted"):
                message = "Veggies production record deleted successfully."
            if request.args.get("saved"):
                message = ("Veggies production saved successfully. "
                           f"Production Date: {_text(request.args.get('saved_date'))}; "
                           f"Assignee: {_text(request.args.get('saved_assignee')) or '—'}; "
                           f"Number of Crops Saved: {_text(request.args.get('saved_crops'))}; "
                           f"Total Quantity Saved: {_text(request.args.get('saved_total'))}.")
            return _render_main(crops, farm_areas, message=message)
        values = request.form.to_dict()
        cleaned, errors = validate_submission(values, crops, farm_areas)
        if errors:
            return _render_main(crops, farm_areas, values=values, errors=errors, message="Please correct the highlighted fields.", bad=True)
        try:
            result = save_submission(cleaned)
        except ValueError as exc:
            return _render_main(crops, farm_areas, values=values, message=str(exc), bad=True)
        except Exception:
            return _render_main(crops, farm_areas, values=values, message="Production could not be saved. No data was committed.", bad=True)
        total = sum((item["quantity"] for item in cleaned["items"]), Decimal("0"))
        return redirect(url_for(
            "veggies_production_basic", saved="1",
            saved_date=cleaned["production_date"].isoformat(),
            saved_assignee=cleaned["assignee"] or "",
            saved_crops=len(cleaned["items"]), saved_total=_quantity_text(total),
        ), code=303)

    @app.get("/veggies-production/crops")
    def veggies_crop_master_page():
        try:
            master_rows = list_crop_master()
            error = ""
        except Exception as exc:
            master_rows = []
            error = f"Could not load Veggies Crop Master: {exc}"
        cards = []
        for crop in master_rows:
            category_options = ''.join(
                f'<option value="{escape(category)}" {"selected" if crop["category"] == category else ""}>{escape(category)}</option>'
                for category in CROP_CATEGORIES
            )
            checked = "checked" if crop["active"] else ""
            cards.append(f'''<section><form method="post" action="/veggies-production/crops/{crop['id']}"><div class="grid"><label>Crop Name<input name="crop_name" value="{escape(crop['crop_name'])}" required></label><label>Category<select name="category">{category_options}</select></label><label>Default Unit<input name="default_unit" value="{escape(_text(crop['default_unit']))}" placeholder="Leave blank until verified"></label><label>Display Order<input type="number" name="display_order" value="{crop['display_order']}" required></label><label>Active<br><input style="width:auto;min-height:auto" type="checkbox" name="active" value="yes" {checked}> Show on new entry form</label></div><div class="actions"><button type="submit">Save Crop</button></div></form></section>''')
        status = f'<p class="status bad">{escape(error)}</p>' if error else ""
        saved = '<p class="status">Veggies Crop Master saved.</p>' if request.args.get("saved") else ""
        return f'''<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Veggies Crop Master</title><style>{BASE_STYLE}</style></head><body><header><h1>Veggies Crop Master</h1></header><main>{status}{saved}<p>Deactivate crops instead of deleting them. Historical production remains linked and visible.</p><div class="actions"><a class="button secondary" href="/veggies-production">Back to Production Entry</a></div>{''.join(cards)}</main></body></html>'''

    @app.post("/veggies-production/crops/<int:crop_id>")
    def veggies_crop_master_update(crop_id):
        try:
            update_crop_master(crop_id, request.form.to_dict())
        except (ValueError, LookupError) as exc:
            return f"Crop could not be saved: {escape(str(exc))}", 400
        except Exception:
            return "Crop could not be saved. No data was committed.", 500
        return redirect("/veggies-production/crops?saved=1", code=303)

    @app.get("/veggies-production/<int:batch_id>")
    def veggies_production_detail(batch_id):
        record = get_record(batch_id)
        if not record:
            return "Production record not found.", 404
        item_rows = ''.join(f'<tr><td>{escape(item["crop_name"])}</td><td>{escape(_quantity_text(item["quantity"]))}</td><td>{escape(_text(item["unit"]) or "—")}</td><td>{escape(_text(item["created_at"]))}</td><td>{escape(_text(item["updated_at"]))}</td></tr>' for item in record["items"])
        return f'''<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Veggies Production Detail</title><style>{BASE_STYLE}</style></head><body><header><h1>Veggies Production Basic</h1></header><main><section><h2>Production record #{batch_id}</h2><div class="grid"><p><b>Production Date</b><br>{escape(_date_text(record['production_date']))}</p><p><b>Farm Area</b><br>{escape(_text(record['farm_area']))}</p><p><b>Assignee</b><br>{escape(_text(record['assignee']) or '—')}</p><p><b>Date of Entry</b><br>{escape(_date_text(record['entry_date']) or '—')}</p><p><b>Created</b><br>{escape(_text(record['created_at']))}</p><p><b>Updated</b><br>{escape(_text(record['updated_at']))}</p></div><p><b>Note</b><br>{escape(_text(record['note']) or '—')}</p><p><b>AI Note</b><br>{escape(_text(record['ai_note']) or '—')}</p><div class="table-wrap"><table><thead><tr><th>Crop</th><th>Quantity</th><th>Unit</th><th>Created time</th><th>Updated time</th></tr></thead><tbody>{item_rows}</tbody></table></div><div class="actions"><a class="button" href="/veggies-production/{batch_id}/edit">Edit</a><a class="button secondary" href="/veggies-production/{batch_id}/pdf">PDF</a><a class="button secondary" href="/veggies-production/{batch_id}/excel">Excel</a><a class="button secondary" href="/veggies-production">Back</a></div><hr><h3>Delete Production Record</h3><p>Imported workbook records are protected. Type record ID <strong>{batch_id}</strong> and provide a reason.</p><form method="post" action="/veggies-production/{batch_id}/delete" onsubmit="return confirm('Permanently delete production record {batch_id}?')"><div class="grid"><label>Record ID<input name="confirmation" required></label><label>Deletion reason<input name="reason" required></label></div><div class="actions"><button type="submit" class="secondary">Delete Record</button></div></form></section></main></body></html>'''

    @app.get("/veggies-production/<int:batch_id>/pdf")
    def veggies_production_pdf(batch_id):
        record = get_record(batch_id)
        if not record:
            return "Production record not found.", 404
        return send_file(_record_pdf(record), mimetype="application/pdf", as_attachment=True,
                         download_name=f"Veggies_Production_{batch_id}.pdf")

    @app.get("/veggies-production/<int:batch_id>/excel")
    def veggies_production_excel(batch_id):
        record = get_record(batch_id)
        if not record:
            return "Production record not found.", 404
        return send_file(_record_excel(record),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"Veggies_Production_{batch_id}.xlsx")

    @app.post("/veggies-production/<int:batch_id>/delete")
    def veggies_production_delete(batch_id):
        try:
            delete_submission(batch_id, request.form.get("confirmation"), request.form.get("reason"))
        except (ValueError, LookupError) as exc:
            return f"Production record could not be deleted: {escape(str(exc))}", 400
        except Exception:
            return "Production record could not be deleted. No data was changed.", 500
        return redirect("/veggies-production?deleted=1", code=303)

    @app.route("/veggies-production/<int:batch_id>/edit", methods=["GET", "POST"])
    def veggies_production_edit(batch_id):
        crops = portal_crops()
        farm_areas = portal_farm_areas()
        record = get_record(batch_id)
        if not record:
            return "Production record not found.", 404
        known_codes = {crop.crop_code for crop in crops}
        for item in record["items"]:
            if item["crop_code"] not in known_codes:
                crops.append(CropDefinition(
                    item["crop_code"], item["crop_name"], item["crop_name"],
                    crop_id=item["crop_id"], default_unit=item["unit"],
                    category=item.get("category") or "Other",
                ))
        original = {
            **record,
            "item_values": {item["crop_code"]: item["quantity"] for item in record["items"]},
        }
        if request.method == "GET":
            values, errors, message = original, {}, ""
        else:
            values = request.form.to_dict()
            cleaned, errors = validate_submission(values, crops, farm_areas)
            if request.form.get("confirm_changes") != "yes":
                errors["confirm_changes"] = "Confirm the changes before saving."
            if not errors:
                try:
                    update_submission(batch_id, cleaned)
                except Exception:
                    return "Production changes could not be saved. No data was committed.", 500
                return redirect(f"/veggies-production/{batch_id}", code=303)
            message = "Please correct the highlighted fields."
        original_text = f"Original: {_date_text(record['production_date'])}; {record.get('assignee') or 'no assignee'}; {len(record['items'])} crops."
        timestamps = f"Original created time: {_text(record.get('created_at'))}. Last updated time: {_text(record.get('updated_at')) or '—'}."
        return f'''<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Edit Veggies Production</title><style>{BASE_STYLE}</style></head><body><header><h1>Veggies Production Basic</h1></header><main><section><h2>Edit production record #{batch_id}</h2><div class="warning"><strong>You are editing an existing production record.</strong><br>{escape(timestamps)}</div><div class="original">{escape(original_text)} The saved record is not changed until you confirm and submit.</div>{f'<p class="status bad">{escape(message)}</p>' if message else ''}{_entry_form(crops, farm_areas, values, errors, f'/veggies-production/{batch_id}/edit', edit=True)}{_error(errors,'confirm_changes')}<div class="actions"><a class="button secondary" href="/veggies-production/{batch_id}">Cancel</a></div></section></main></body></html>'''
