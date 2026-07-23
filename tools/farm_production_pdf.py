from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _quantity(value):
    return f"{float(value or 0):,.2f}"


def write_farm_production_pdf(data, path):
    """Render the Farm Production report from the same canonical API payload as the UI."""
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title="BigShot Farm Production",
    )
    story = [
        Paragraph("BigShot Farm Production", styles["Title"]),
        Paragraph(
            f"Period: {data.get('start_date')} to {data.get('end_date')} · Daily timeline",
            styles["Normal"],
        ),
        Spacer(1, 5 * mm),
    ]
    summary = data.get("summary") or {}
    kpis = [
        ["Total Production", "Production Days", "Active Crops", "Top Field", "Top Crop"],
        [
            f"{_quantity(summary.get('total_production'))} {summary.get('total_unit') or 'Unspecified'}",
            str(summary.get("production_days") or 0),
            str(summary.get("active_crops") or 0),
            str(summary.get("top_field") or "—"),
            str(summary.get("top_crop") or "—"),
        ],
    ]
    story.append(Table(kpis, repeatRows=1, colWidths=[51 * mm] * 5))
    story.append(Spacer(1, 6 * mm))

    crop_names = sorted({
        crop for row in data.get("daily_stacked") or [] for crop in (row.get("crops") or {})
    })
    crop_units = {
        crop: next((
            row.get("crop_units", {}).get(crop)
            for row in data.get("daily_stacked") or []
            if row.get("crop_units", {}).get(crop)
        ), "Unspecified")
        for crop in crop_names
    }
    daily_rows = [["Date", *[f"{crop} ({crop_units[crop]})" for crop in crop_names], "Total"]]
    for row in data.get("daily_stacked") or []:
        crops = row.get("crops") or {}
        daily_rows.append([
            row.get("label") or row.get("period"),
            *[_quantity(crops.get(crop)) for crop in crop_names],
            (f"{_quantity(row.get('total'))} {row.get('total_unit')}"
             if row.get("total_unit") != "Mixed units" else "Mixed units"),
        ])
    daily = Table(daily_rows, repeatRows=1)
    daily.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#176b5d")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d7dfdb")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
    ]))
    story.extend([Paragraph("Daily Production", styles["Heading2"]), daily, Spacer(1, 6 * mm)])

    records = [["Date", "Crop", "Field", "Quantity", "Unit"]]
    for row in data.get("combined_rows") or []:
        records.append([
            str(row.get("production_date") or ""),
            str(row.get("crop_name") or ""),
            str(row.get("farm_area") or ""),
            _quantity(row.get("quantity")),
            str(row.get("unit") or ""),
        ])
    record_table = Table(records, repeatRows=1)
    record_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#176b5d")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d7dfdb")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (3, 1), (3, -1), "RIGHT"),
    ]))
    story.extend([Paragraph("Production Records", styles["Heading2"]), record_table])
    document.build(story)


def write_farm_production_excel(data, path):
    """Write the resolved canonical Farm Production payload without querying data again."""
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary = data.get("summary") or {}
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append([
        "Total Production",
        (f"{_quantity(summary.get('total_production'))} {summary.get('total_unit') or 'Unspecified'}"
         if summary.get("total_unit") != "Mixed units" else "Mixed units"),
    ])
    summary_sheet.append(["Production Days", summary.get("production_days") or 0])
    summary_sheet.append(["Active Crops", summary.get("active_crops") or 0])
    summary_sheet.append(["Top Crop", summary.get("top_crop") or "—"])
    summary_sheet.append(["Top Crop Unit", summary.get("top_crop_unit") or "Unspecified"])

    records = workbook.create_sheet("Production Records")
    records.append(["Date", "Crop Code", "Crop", "Field", "Quantity", "Unit"])
    for row in data.get("combined_rows") or []:
        records.append([
            row.get("production_date"),
            row.get("crop_code") or "",
            row.get("crop_name") or "",
            row.get("farm_area") or "",
            float(row.get("quantity") or 0),
            row.get("unit") or "Unspecified",
        ])
    for sheet in (summary_sheet, records):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="176B5D")
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(
                42, max(12, max(len(str(cell.value or "")) for cell in column) + 2)
            )
    workbook.save(path)
