"""Parsing, validation, normalization, and persistence for vegetable production."""

from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, BinaryIO, Iterable

import psycopg2.extras
from openpyxl import load_workbook
from openpyxl.utils.datetime import WINDOWS_EPOCH, from_excel

from tools.formula_engine import _connect


STANDARD_CROPS = [
    ("ZUCCHINI", "Zucchini", "Zucchini"),
    ("CHERRY_TOMATO", "Cherry Tomato", "Cherry Tomato"),
    ("ROSEMARY", "Rosemary", "Rosemary"),
    ("ROMAINE", "Romaine Lettuce", "Romaine"),
    ("ICEBERG_LETTUCE", "Iceberg Lettuce", "Iceberg"),
    ("GREEN_OAK_LETTUCE", "Green Oak Lettuce", "Green Oak"),
    ("RED_OAK_LETTUCE", "Red Oak Lettuce", "Red Oak"),
    ("GREEN_LOLLO_LETTUCE", "Green Lollo Lettuce", "Green lollo"),
    ("RED_LOLLO_LETTUCE", "Red Lollo Lettuce", "Red lollo"),
    ("DAIKON", "Daikon", "Daikon"),
    ("LEEK", "Leek", "Leek"),
    ("LONG_CHILI", "Long Chili", "Long Chilli"),
    ("BELL_PEPPER", "Bell Pepper", "Bell Pepper"),
    ("RED_CABBAGE", "Red Cabbage", "Red Cabbage"),
    ("JAPANESE_CUCUMBER", "Japanese Cucumber", "Japanese Cucumber"),
    ("BEETROOT", "Beetroot", "Beet root"),
    ("CARROT", "Carrot", "Carrot"),
    ("EDAMAME", "Edamame", "Edamame"),
    ("SWISS_CHARD", "Swiss Chard", "Swiss Chert"),
    ("ROCKET", "Rocket", "Rocket"),
    ("PARSLEY", "Parsley", "Persley"),
    ("KALE", "Kale", "Kale"),
    ("RED_RADISH", "Red Radish", "Red Radish"),
    ("JAPANESE_PUMPKIN", "Japanese Pumpkin", "Japanese Pumpkin"),
    ("THYME", "Thyme", "Thyme"),
    ("GREEN_PERILLA", "Green Perilla", "Green Perilla"),
    ("LONG_SWEET_PEPPER", "Long Sweet Pepper", "Long Sweet Pepper"),
    ("FANCY_TOMATO", "Fancy Tomato", "Fancy Tomato"),
    ("ASPARAGUS", "Asparagus", "Asparagus"),
    ("BRUSSELS_SPROUTS", "Brussels Sprouts", "Brussel Sprout"),
    ("BASIL", "Basil", "Basil"),
    ("FENNEL_BULB", "Fennel Bulb", "Funnel Bulb"),
]

CROP_CATEGORIES = (
    "Leafy Vegetables",
    "Fruit Vegetables",
    "Root and Bulb Vegetables",
    "Herbs and Specialty Crops",
    "Legumes and Others",
    "Other",
)

CROP_CATEGORY_BY_CODE = {
    "ROMAINE": "Leafy Vegetables", "ICEBERG_LETTUCE": "Leafy Vegetables",
    "GREEN_OAK_LETTUCE": "Leafy Vegetables", "RED_OAK_LETTUCE": "Leafy Vegetables",
    "GREEN_LOLLO_LETTUCE": "Leafy Vegetables", "RED_LOLLO_LETTUCE": "Leafy Vegetables",
    "SWISS_CHARD": "Leafy Vegetables", "ROCKET": "Leafy Vegetables", "KALE": "Leafy Vegetables",
    "ZUCCHINI": "Fruit Vegetables", "CHERRY_TOMATO": "Fruit Vegetables",
    "LONG_CHILI": "Fruit Vegetables", "BELL_PEPPER": "Fruit Vegetables",
    "JAPANESE_CUCUMBER": "Fruit Vegetables", "LONG_SWEET_PEPPER": "Fruit Vegetables",
    "FANCY_TOMATO": "Fruit Vegetables", "JAPANESE_PUMPKIN": "Fruit Vegetables",
    "DAIKON": "Root and Bulb Vegetables", "LEEK": "Root and Bulb Vegetables",
    "RED_CABBAGE": "Root and Bulb Vegetables", "BEETROOT": "Root and Bulb Vegetables",
    "CARROT": "Root and Bulb Vegetables", "RED_RADISH": "Root and Bulb Vegetables",
    "FENNEL_BULB": "Root and Bulb Vegetables",
    "ROSEMARY": "Herbs and Specialty Crops", "PARSLEY": "Herbs and Specialty Crops",
    "THYME": "Herbs and Specialty Crops", "GREEN_PERILLA": "Herbs and Specialty Crops",
    "BASIL": "Herbs and Specialty Crops", "ASPARAGUS": "Herbs and Specialty Crops",
    "EDAMAME": "Legumes and Others", "BRUSSELS_SPROUTS": "Legumes and Others",
}

ADMIN_HEADERS = {
    "date": "production_date",
    "production date": "production_date",
    "asignee": "assignee",
    "assignee": "assignee",
    "note": "note",
    "ai note": "ai_note",
    "date of entry": "entry_date",
}


def normalize_header(value: Any) -> str:
    """Return a stable comparison key while retaining source headers elsewhere."""
    text = str(value or "").strip().lower()
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", text)).strip()


def _schema() -> str:
    """Veggies tables are Business OS public-schema objects."""
    return "public"


def _ref(name: str) -> str:
    return f'"{_schema()}"."{name}"'


def parse_production_date(value: Any, field: str = "Production Date", epoch=WINDOWS_EPOCH) -> date | None:
    """Parse supported Excel and text dates into a real ``date`` value."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        try:
            parsed = from_excel(float(value), epoch=epoch)
        except Exception as exc:
            raise ValueError(f"{field} contains an invalid Excel serial date: {value}") from exc
        return parsed.date() if isinstance(parsed, datetime) else parsed
    text = str(value).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(
        f"{field} must be an Excel date or use YYYY/MM/DD, YYYY-MM-DD, or DD/MM/YYYY"
    )


def parse_quantity(value: Any, crop_name: str) -> Decimal | None:
    """Parse a crop quantity, preserving blank versus explicit zero."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, bool):
        raise ValueError(f'{crop_name} quantity "{value}" is not numeric')
    try:
        quantity = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f'{crop_name} quantity "{value}" is not numeric') from exc
    if not quantity.is_finite():
        raise ValueError(f'{crop_name} quantity "{value}" is not numeric')
    if quantity < 0:
        raise ValueError(f"{crop_name} quantity cannot be negative")
    return quantity


def format_quantity(value: Any, blank: str = "") -> str:
    """Format a production quantity for display without changing its stored value."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return blank
    quantity = Decimal(str(value).replace(",", "").strip())
    return format(quantity.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), ".2f")


@dataclass(frozen=True)
class CropDefinition:
    crop_code: str
    crop_name: str
    source_header: str
    crop_id: int | None = None
    default_unit: str | None = None
    category: str = "Other"


@dataclass(frozen=True)
class RowError:
    row_number: int
    column: str
    message: str

    def as_dict(self) -> dict[str, Any]:
        return {
            "row_number": self.row_number,
            "column": self.column,
            "message": self.message,
        }


@dataclass
class NormalizedRow:
    row_number: int
    production_date: date
    assignee: str | None
    note: str | None
    ai_note: str | None
    entry_date: date | None
    items: list[dict[str, Any]]
    row_hash: str

    def as_preview_dict(self) -> dict[str, Any]:
        return {
            "row_number": self.row_number,
            "production_date": self.production_date.isoformat(),
            "assignee": self.assignee,
            "entry_date": self.entry_date.isoformat() if self.entry_date else None,
            "crop_count": len(self.items),
            "items": [
                {
                    **item,
                    "quantity": str(item["quantity"]),
                }
                for item in self.items
            ],
            "row_hash": self.row_hash,
        }


@dataclass
class WorkbookPreview:
    filename: str
    workbook_name: str
    file_hash: str
    total_source_rows: int
    valid_rows: list[NormalizedRow] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)
    duplicate_rows: list[int] = field(default_factory=list)
    unknown_headers: list[str] = field(default_factory=list)

    @property
    def accepted_rows(self) -> int:
        return len(self.valid_rows)

    @property
    def rejected_rows(self) -> int:
        return len({error.row_number for error in self.errors})

    @property
    def item_count(self) -> int:
        return sum(len(row.items) for row in self.valid_rows)

    def as_dict(self, include_rows: bool = True) -> dict[str, Any]:
        payload = {
            "filename": self.filename,
            "workbook_name": self.workbook_name,
            "file_hash": self.file_hash,
            "total_source_rows": self.total_source_rows,
            "accepted_rows": self.accepted_rows,
            "rejected_rows": self.rejected_rows,
            "normalized_items": self.item_count,
            "duplicate_rows": self.duplicate_rows,
            "unknown_headers": self.unknown_headers,
            "errors": [error.as_dict() for error in self.errors],
        }
        if include_rows:
            payload["rows"] = [row.as_preview_dict() for row in self.valid_rows]
        return payload


def default_crop_definitions() -> list[CropDefinition]:
    return [
        CropDefinition(code, name, source, category=CROP_CATEGORY_BY_CODE.get(code, "Other"))
        for code, name, source in STANDARD_CROPS
    ]


def load_crop_definitions(connection=None) -> list[CropDefinition]:
    """Load active crop names and aliases, falling back to migration seed definitions."""
    owns_connection = connection is None
    if connection is None:
        try:
            connection = _connect()
        except Exception:
            return default_crop_definitions()
    try:
        with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(
                f"""
                SELECT crop.id AS crop_id, crop.crop_code, crop.crop_name,
                       crop.default_unit, crop.category,
                       COALESCE(alias.source_header, crop.crop_name) AS source_header
                FROM {_ref('veggies_crop_master')} crop
                LEFT JOIN {_ref('veggies_crop_alias')} alias ON alias.crop_id = crop.id
                WHERE crop.active = TRUE
                ORDER BY crop.display_order, crop.crop_name, alias.id
                """
            )
            return [CropDefinition(**dict(row)) for row in cursor.fetchall()]
    except Exception:
        if owns_connection:
            connection.rollback()
        return default_crop_definitions()
    finally:
        if owns_connection:
            connection.close()


def crop_header_map(crops: Iterable[CropDefinition]) -> dict[str, CropDefinition]:
    result: dict[str, CropDefinition] = {}
    for crop in crops:
        result[normalize_header(crop.crop_code)] = crop
        result[normalize_header(crop.crop_name)] = crop
        result[normalize_header(crop.source_header)] = crop
    return result


def _clean_optional(value: Any) -> str | None:
    value = str(value or "").strip()
    return value or None


def _canonical_decimal(value: Decimal) -> str:
    return format(value.normalize(), "f") if value else "0"


def _row_hash(production_date: date, assignee: str | None, entry_date: date | None,
              note: str | None, ai_note: str | None, items: list[dict[str, Any]]) -> str:
    payload = {
        "production_date": production_date.isoformat(),
        "assignee": assignee or "",
        "entry_date": entry_date.isoformat() if entry_date else "",
        "note": note or "",
        "ai_note": ai_note or "",
        "items": sorted(
            (item["crop_code"], _canonical_decimal(item["quantity"])) for item in items
        ),
    }
    return hashlib.sha256(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()


def _file_bytes(source: str | Path | BinaryIO) -> tuple[bytes, str]:
    if hasattr(source, "read"):
        data = source.read()
        if hasattr(source, "seek"):
            source.seek(0)
        return data, getattr(source, "name", "veggies-production.xlsx")
    path = Path(source)
    return path.read_bytes(), path.name


def parse_veggies_workbook(source: str | Path | BinaryIO,
                           crops: Iterable[CropDefinition] | None = None) -> WorkbookPreview:
    """Parse a wide workbook into validated normalized batch/item records."""
    data, filename = _file_bytes(source)
    from io import BytesIO

    workbook = load_workbook(BytesIO(data), data_only=True, read_only=False)
    sheet = workbook["Data Entry"] if "Data Entry" in workbook.sheetnames else workbook.active
    headers = [cell.value for cell in sheet[1]]
    crops = list(crops or default_crop_definitions())
    crop_map = crop_header_map(crops)
    crop_columns: dict[int, CropDefinition] = {}
    admin_columns: dict[str, int] = {}
    unknown_headers: list[str] = []

    for column_index, header in enumerate(headers, start=1):
        key = normalize_header(header)
        if not key:
            continue
        if key in ADMIN_HEADERS:
            admin_columns[ADMIN_HEADERS[key]] = column_index
        elif key in crop_map:
            crop_columns[column_index] = crop_map[key]
        else:
            unknown_headers.append(str(header).strip())

    preview = WorkbookPreview(
        filename=filename,
        workbook_name=sheet.title,
        file_hash=hashlib.sha256(data).hexdigest(),
        total_source_rows=0,
        unknown_headers=unknown_headers,
    )
    if "production_date" not in admin_columns:
        preview.errors.append(RowError(1, "Production Date", "Required header is missing"))
        return preview
    if unknown_headers:
        for header in unknown_headers:
            preview.errors.append(RowError(1, header, "Crop header is not recognized"))
        return preview

    seen_hashes: dict[str, int] = {}
    for row_number in range(2, sheet.max_row + 1):
        values = [sheet.cell(row_number, col).value for col in range(1, sheet.max_column + 1)]
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        preview.total_source_rows += 1
        row_errors: list[RowError] = []

        raw_date = sheet.cell(row_number, admin_columns["production_date"]).value
        try:
            production_date = parse_production_date(raw_date)
            if production_date is None:
                raise ValueError("Production Date is required")
        except ValueError as exc:
            production_date = None
            row_errors.append(RowError(row_number, "Production Date", str(exc)))

        entry_date = None
        if "entry_date" in admin_columns:
            raw_entry_date = sheet.cell(row_number, admin_columns["entry_date"]).value
            try:
                entry_date = parse_production_date(raw_entry_date, "Date of Entry")
            except ValueError as exc:
                row_errors.append(RowError(row_number, "Date of Entry", str(exc)))

        items = []
        for column_index, crop in crop_columns.items():
            raw_quantity = sheet.cell(row_number, column_index).value
            try:
                quantity = parse_quantity(raw_quantity, crop.crop_name)
            except ValueError as exc:
                row_errors.append(RowError(row_number, str(headers[column_index - 1]), str(exc)))
                continue
            if quantity is not None:
                items.append({
                    "crop_id": crop.crop_id,
                    "crop_code": crop.crop_code,
                    "crop_name": crop.crop_name,
                    "quantity": quantity,
                    "unit": crop.default_unit,
                })

        if not items:
            row_errors.append(RowError(row_number, "Crop quantities", "At least one crop quantity is required"))

        assignee = _clean_optional(
            sheet.cell(row_number, admin_columns["assignee"]).value
            if "assignee" in admin_columns else None
        )
        note = _clean_optional(
            sheet.cell(row_number, admin_columns["note"]).value
            if "note" in admin_columns else None
        )
        ai_note = _clean_optional(
            sheet.cell(row_number, admin_columns["ai_note"]).value
            if "ai_note" in admin_columns else None
        )

        if row_errors:
            preview.errors.extend(row_errors)
            continue

        row_hash = _row_hash(production_date, assignee, entry_date, note, ai_note, items)
        if row_hash in seen_hashes:
            preview.duplicate_rows.append(row_number)
            preview.errors.append(RowError(
                row_number,
                "Row",
                f"Duplicate of source row {seen_hashes[row_hash]}",
            ))
            continue
        seen_hashes[row_hash] = row_number
        preview.valid_rows.append(NormalizedRow(
            row_number=row_number,
            production_date=production_date,
            assignee=assignee,
            note=note,
            ai_note=ai_note,
            entry_date=entry_date,
            items=items,
            row_hash=row_hash,
        ))
    return preview


def import_veggies_preview(preview: WorkbookPreview, imported_by: str | None = None,
                           connection=None) -> dict[str, Any]:
    """Insert valid preview rows atomically; never overwrite existing production."""
    owns_connection = connection is None
    connection = connection or _connect()
    try:
        with connection.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(
                f"""
                SELECT id FROM {_ref('veggies_production_imports')}
                WHERE file_hash = %s AND status IN ('completed', 'completed_with_errors')
                LIMIT 1
                """,
                (preview.file_hash,),
            )
            existing = cursor.fetchone()
            if existing:
                connection.rollback()
                return {
                    "ok": True,
                    "already_exists": True,
                    "import_id": existing["id"],
                    "accepted_rows": 0,
                    "rejected_rows": preview.rejected_rows,
                    "created_batches": 0,
                    "created_items": 0,
                    "duplicate_rows": preview.total_source_rows,
                }

            status = "completed_with_errors" if preview.errors else "completed"
            cursor.execute(
                f"""
                INSERT INTO {_ref('veggies_production_imports')}
                  (filename, workbook_name, file_hash, imported_by, total_source_rows,
                   accepted_rows, rejected_rows, created_batches, created_items,
                   duplicate_rows, status, completed_at, error_summary)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), %s::jsonb)
                RETURNING id
                """,
                (
                    preview.filename,
                    preview.workbook_name,
                    preview.file_hash,
                    imported_by,
                    preview.total_source_rows,
                    preview.accepted_rows,
                    preview.rejected_rows,
                    preview.accepted_rows,
                    preview.item_count,
                    len(preview.duplicate_rows),
                    status,
                    json.dumps([error.as_dict() for error in preview.errors]),
                ),
            )
            import_id = cursor.fetchone()["id"]

            crop_ids = {}
            cursor.execute(
                f"SELECT id, crop_code FROM {_ref('veggies_crop_master')} WHERE active = TRUE"
            )
            for crop in cursor.fetchall():
                crop_ids[crop["crop_code"]] = crop["id"]
            cursor.execute(
                f"SELECT id FROM {_ref('veggies_farm_area_master')} WHERE area_code = 'HOME_FARM' AND active = TRUE"
            )
            home_farm = cursor.fetchone()
            if not home_farm:
                raise ValueError("Active Home Farm is missing from Farm Area Master")

            for row in preview.valid_rows:
                cursor.execute(
                    f"""
                    INSERT INTO {_ref('veggies_production_batches')}
                      (production_date, farm_area_id, assignee, note, ai_note, entry_date, source_file,
                       source_workbook, import_id, source_row_number, source_row_hash, created_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (
                        row.production_date, home_farm["id"], row.assignee, row.note, row.ai_note, row.entry_date,
                        preview.filename, preview.workbook_name, import_id, row.row_number,
                        row.row_hash, imported_by,
                    ),
                )
                batch_id = cursor.fetchone()["id"]
                for item in row.items:
                    crop_id = item.get("crop_id") or crop_ids.get(item["crop_code"])
                    if not crop_id:
                        raise ValueError(f"Active crop is missing from master: {item['crop_name']}")
                    cursor.execute(
                        f"""
                        INSERT INTO {_ref('veggies_production_items')}
                          (production_batch_id, crop_id, quantity, unit)
                        VALUES (%s, %s, %s, %s)
                        """,
                        (batch_id, crop_id, item["quantity"], item.get("unit")),
                    )
        connection.commit()
        return {
            "ok": True,
            "already_exists": False,
            "import_id": import_id,
            "accepted_rows": preview.accepted_rows,
            "rejected_rows": preview.rejected_rows,
            "created_batches": preview.accepted_rows,
            "created_items": preview.item_count,
            "duplicate_rows": len(preview.duplicate_rows),
        }
    except Exception:
        connection.rollback()
        raise
    finally:
        if owns_connection:
            connection.close()
